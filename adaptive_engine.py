from __future__ import annotations

import io
import re
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

import bom_engine as legacy_rsp
import hlg_engine


class AdaptiveBOMError(ValueError):
    """Raised when automatic detection is not reliable enough to generate."""


Source = bytes | bytearray | str | Path | BinaryIO

SECTION_RE = re.compile(
    r"^\s*(?:SMT|ASSY|DIP)\s*([A-Z][A-Z0-9-]*)\s*$", re.IGNORECASE
)
WORK_ORDER_RE = re.compile(r"^W[A-Z0-9-]{5,}$", re.IGNORECASE)
MODEL_ENTRY_RE = re.compile(
    r"(?P<number>\d+)\s*、\s*"
    r"(?P<name>[A-Z]{2}-[A-Z0-9][A-Z0-9-]*)\s*"
    r"\((?P<quantity>[\d,]+)\)\s*"
    r"\((?P<secondary>[\d,]+)\)",
    re.IGNORECASE,
)
SEMI_RE = re.compile(r"^9\S+-P$", re.IGNORECASE)
PART_RE = re.compile(r"^[A-Za-z0-9][^\s=]{2,}$")
POSITION_RE = re.compile(
    r"^(?:PCB|[A-Z]{1,8}\d+[A-Z0-9_*-]*)$", re.IGNORECASE
)


def _has_pcb(positions: Sequence[str]) -> bool:
    return any(position.upper().startswith("PCB") for position in positions)


@dataclass(frozen=True)
class PartSpec:
    part_no: str
    specification: str


@dataclass(frozen=True)
class ModelDefinition:
    number: int
    name: str
    quantity: int | None = None
    secondary_quantity: int | None = None


@dataclass(frozen=True)
class DiffRule:
    section: str
    position: str
    part_no: str
    specification: str
    models: frozenset[int]
    source_order: int
    source_row: int


@dataclass
class DifferenceGroup:
    work_order: str
    models: OrderedDict[int, ModelDefinition] = field(default_factory=OrderedDict)
    rules_by_section: OrderedDict[str, list[DiffRule]] = field(
        default_factory=OrderedDict
    )
    source_start: int = 0
    source_end: int = 0

    @property
    def section_order(self) -> tuple[str, ...]:
        return tuple(self.rules_by_section)

    def sections_for_model(self, model: int) -> tuple[str, ...]:
        return tuple(
            section
            for section, rules in self.rules_by_section.items()
            if any(model in rule.models for rule in rules)
        )

    @property
    def family_tokens(self) -> set[tuple[str, str]]:
        tokens: set[tuple[str, str]] = set()
        for definition in self.models.values():
            tokens.update(_family_tokens(definition.name))
        return tokens


@dataclass
class SourceItem:
    part_no: str
    specification: str
    positions: list[str]
    source_order: float
    declared_quantity: int | None = None


@dataclass
class BaselineRoot:
    semi_finished: str
    process: str
    source_items: list[SourceItem]
    source_row: int
    batch_multiplier: int = 1

    @property
    def position_set(self) -> set[str]:
        return {
            position
            for item in self.source_items
            for position in item.positions
        }

    @property
    def identity_text(self) -> str:
        pcb_items = [
            item for item in self.source_items if "PCB" in item.positions
        ]
        parts = [self.semi_finished]
        for item in pcb_items or self.source_items[:1]:
            parts.extend((item.part_no, item.specification))
        return " ".join(parts).upper()


@dataclass
class BaselineGroup:
    index: int
    roots: list[BaselineRoot]
    source_start: int
    source_end: int

    @property
    def family_tokens(self) -> set[tuple[str, str]]:
        tokens: set[tuple[str, str]] = set()
        for root in self.roots:
            tokens.update(_family_tokens(root.identity_text))
        return tokens


@dataclass(frozen=True)
class SectionMapping:
    section: str
    baseline_root_index: int | None
    method: str
    confidence: int


@dataclass
class PairedGroup:
    difference: DifferenceGroup
    baseline: BaselineGroup
    mappings: dict[str, SectionMapping]
    pairing_score: int


@dataclass(frozen=True)
class TargetPlan:
    key: str
    work_order: str
    model: int
    model_name: str
    quantity: int | None
    sections: tuple[str, ...]
    baseline_sections: tuple[str, ...]
    inferred_sections: tuple[str, ...]


@dataclass
class AdaptiveAnalysis:
    mode: str
    targets: list[TargetPlan]
    warnings: list[str]
    paired_groups: list[PairedGroup] = field(default_factory=list)
    legacy_details: dict[str, object] | None = None

    @property
    def expected_sheet_count(self) -> int:
        return sum(len(target.sections) for target in self.targets)

    @property
    def matrix(self) -> list[dict[str, object]]:
        return [
            {
                "工單": target.work_order if target.work_order != "AUTO" else "單一系列",
                "機種": f"第{target.model}機種",
                "機種名稱": target.model_name,
                "自動辨識版別": "/".join(target.sections),
                "工作表數": len(target.sections),
                "基準來源": "/".join(target.baseline_sections) or "差異表建立",
                "需推導版別": "/".join(target.inferred_sections),
            }
            for target in self.targets
        ]


@dataclass
class OutputItem:
    part_no: str
    specification: str
    quantity: int
    positions: list[str]
    source_order: float
    changed_positions: set[str] = field(default_factory=set)


@dataclass
class GeneratedSheet:
    target_key: str
    work_order_group: str
    model: int
    section: str
    sheet_name: str
    model_name: str
    work_order: str
    order_quantity: int | None
    semi_finished: str
    process: str
    items: list[OutputItem]

    @property
    def total_quantity(self) -> int:
        return sum(
            item.quantity for item in self.items if not _has_pcb(item.positions)
        )


@dataclass
class GenerationResult:
    workbook_bytes: bytes
    sheets: list[GeneratedSheet]
    warnings: list[str]
    mode: str
    target_keys: tuple[str, ...]


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _compact(value: object) -> str:
    return re.sub(r"\s+", "", _clean(value))


def _load(source: Source):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    return load_workbook(source, data_only=True)


def _part_key(value: str) -> str:
    return re.sub(r"\s+", "", value).upper()


def _spec_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().casefold()


def _identity(part_no: str, specification: str) -> tuple[str, str]:
    return _part_key(part_no), _spec_key(specification)


def _is_number(value: object) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    try:
        float(_clean(value).replace(",", ""))
    except ValueError:
        return False
    return bool(_clean(value))


def _as_int(value: object) -> int | None:
    if not _is_number(value):
        return None
    return int(float(_clean(value).replace(",", "")))


def split_positions(value: object) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    text = re.sub(r"[,，、;；\n\r\t]+", " ", text)
    result: list[str] = []
    for raw in re.split(r"\s+", text):
        token = raw.strip().upper()
        if token and POSITION_RE.fullmatch(token) and token not in result:
            result.append(token)
    return result


def _looks_like_part(value: str) -> bool:
    text = _clean(value)
    if not text or " " in text or "=" in text:
        return False
    if text.upper() in {"PCB", "END", "SMT", "ASSY", "DIP"}:
        return False
    return bool(PART_RE.fullmatch(text)) and any(char.isdigit() for char in text)


def _parse_membership(value: object) -> frozenset[int]:
    if value is None or isinstance(value, bool):
        return frozenset()
    if isinstance(value, int):
        return frozenset({value}) if 0 < value < 100 else frozenset()
    if isinstance(value, float) and value.is_integer():
        number = int(value)
        return frozenset({number}) if 0 < number < 100 else frozenset()
    text = _clean(value)
    if not text or not re.fullmatch(r"[\d\s,，;/；]+", text):
        return frozenset()
    return frozenset(
        int(match)
        for match in re.findall(r"(?<!\d)\d{1,2}(?!\d)", text)
        if 0 < int(match) < 100
    )


def _family_tokens(value: str) -> set[tuple[str, str]]:
    upper = _clean(value).upper()
    return {
        (prefix, number)
        for prefix, number in re.findall(r"([A-Z]{2,})[-_ ]*(\d{2,5})", upper)
        if prefix not in {"SMD", "PCB", "FR"}
    }


def _model_entries(value: object) -> OrderedDict[int, ModelDefinition]:
    result: OrderedDict[int, ModelDefinition] = OrderedDict()
    for match in MODEL_ENTRY_RE.finditer(_clean(value)):
        number = int(match.group("number"))
        result[number] = ModelDefinition(
            number=number,
            name=match.group("name").strip(),
            quantity=int(match.group("quantity").replace(",", "")),
            secondary_quantity=int(match.group("secondary").replace(",", "")),
        )
    return result


def parse_difference_groups(source: Source) -> list[DifferenceGroup]:
    workbook = _load(source)
    groups: OrderedDict[str, DifferenceGroup] = OrderedDict()
    generic_group: DifferenceGroup | None = None

    for worksheet in workbook.worksheets:
        current_group: DifferenceGroup | None = None
        current_section: str | None = None
        sequence_by_group: defaultdict[str, int] = defaultdict(int)

        for row_number, row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            values = list(row)
            texts = [_clean(value) for value in values]
            compact_texts = [_compact(value) for value in values]

            work_order = next(
                (
                    text.upper()
                    for text in texts
                    if WORK_ORDER_RE.fullmatch(text)
                ),
                None,
            )
            is_difference_title = any("差異表" in text for text in compact_texts)
            if work_order and is_difference_title:
                previous_key = current_group.work_order if current_group else None
                current_group = groups.setdefault(
                    work_order,
                    DifferenceGroup(work_order=work_order, source_start=row_number),
                )
                current_group.source_end = row_number
                if previous_key != work_order:
                    current_section = None
                continue

            if any("比較機種" in text for text in compact_texts):
                entries: OrderedDict[int, ModelDefinition] = OrderedDict()
                for value in values:
                    entries.update(_model_entries(value))
                if entries:
                    if current_group is None:
                        generic_group = generic_group or DifferenceGroup(
                            work_order="AUTO", source_start=row_number
                        )
                        current_group = generic_group
                    for number, definition in entries.items():
                        existing = current_group.models.get(number)
                        if existing and existing != definition:
                            raise AdaptiveBOMError(
                                f"{current_group.work_order} 的第{number}機種名稱在不同頁不一致。"
                            )
                        current_group.models[number] = definition
                continue

            section_match = next(
                (
                    match
                    for text in texts
                    if text and (match := SECTION_RE.fullmatch(text))
                ),
                None,
            )
            if section_match:
                if current_group is None:
                    generic_group = generic_group or DifferenceGroup(
                        work_order="AUTO", source_start=row_number
                    )
                    current_group = generic_group
                current_section = section_match.group(1).upper()
                current_group.rules_by_section.setdefault(current_section, [])
                current_group.source_end = row_number
                continue

            if current_group is None or current_section is None:
                continue
            nonempty = [(index, text) for index, text in enumerate(texts) if text]
            if len(nonempty) < 3:
                continue
            _, position = nonempty[0]
            _, part_no = nonempty[1]
            spec_index, specification = nonempty[2]
            position = position.upper()
            if not POSITION_RE.fullmatch(position) or not _looks_like_part(part_no):
                continue

            models: set[int] = set()
            for value in values[spec_index + 1 :]:
                models.update(_parse_membership(value))
            if not models:
                continue

            key = current_group.work_order
            sequence_by_group[key] += 1
            current_group.rules_by_section[current_section].append(
                DiffRule(
                    section=current_section,
                    position=position,
                    part_no=part_no,
                    specification=specification,
                    models=frozenset(models),
                    source_order=sequence_by_group[key],
                    source_row=row_number,
                )
            )
            current_group.source_end = row_number

    if generic_group and generic_group.rules_by_section:
        groups[generic_group.work_order] = generic_group
    populated = [
        group
        for group in groups.values()
        if any(group.rules_by_section.values())
    ]
    if not populated:
        raise AdaptiveBOMError("差異表找不到可辨識的製程區、位置、品號與機種數字。")

    for group in populated:
        found_models = sorted(
            {
                model
                for rules in group.rules_by_section.values()
                for rule in rules
                for model in rule.models
            }
        )
        if not group.models:
            for model in found_models:
                group.models[model] = ModelDefinition(
                    number=model,
                    name=f"第{model}機種",
                )
        unknown = [model for model in found_models if model not in group.models]
        if unknown:
            raise AdaptiveBOMError(
                f"{group.work_order} 的差異列使用未列在比較機種中的編號："
                + ", ".join(map(str, unknown))
            )
    return populated


def _is_baseline_header(values: Sequence[object]) -> bool:
    texts = [_compact(value) for value in values]
    return any("製程段" in text for text in texts) and any(
        "半成品" in text for text in texts
    )


def _is_difference_header(values: Sequence[object]) -> bool:
    texts = [_compact(value) for value in values]
    has_process = any(text == "製程" for text in texts)
    has_position = any("位置" in text for text in texts)
    return has_process and has_position and not any("半成品" in text for text in texts)


def _parse_source_item(values: Sequence[object], source_row: int) -> SourceItem | None:
    texts = [_clean(value) for value in values]
    candidates: list[tuple[int, int, list[str]]] = []
    for index in range(1, len(values)):
        positions = split_positions(values[index])
        quantity = _as_int(values[index - 1])
        if positions and quantity is not None:
            candidates.append((len(positions), index, positions))
    if not candidates:
        return None
    _, position_index, positions = max(candidates, key=lambda value: value[0])
    quantity = _as_int(values[position_index - 1])

    semi_indices = {
        index for index, text in enumerate(texts) if SEMI_RE.fullmatch(text)
    }
    part_candidates = [
        (index, text)
        for index, text in enumerate(texts[: position_index - 1])
        if index not in semi_indices and _looks_like_part(text)
    ]
    if not part_candidates:
        return None
    part_index, part_no = part_candidates[0]

    specification = ""
    for index in range(part_index + 1, position_index - 1):
        candidate = texts[index]
        if not candidate or _is_number(candidate) or set(candidate) <= {"=", "."}:
            continue
        if candidate in {"NEW", "2020"}:
            continue
        specification = candidate
        break
    if not specification:
        return None
    return SourceItem(
        part_no=part_no,
        specification=specification,
        positions=positions,
        source_order=float(source_row),
        declared_quantity=quantity,
    )


def parse_baseline_groups(source: Source) -> tuple[list[BaselineGroup], list[str]]:
    workbook = _load(source)
    row_groups: list[tuple[int, int, list[tuple[int, tuple[object, ...]]]]] = []
    global_offset = 0

    for worksheet in workbook.worksheets:
        state: str | None = None
        current_rows: list[tuple[int, tuple[object, ...]]] | None = None
        current_start = 0
        for local_row, values in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            row_number = global_offset + local_row
            if _is_baseline_header(values):
                if state != "baseline":
                    current_rows = []
                    current_start = row_number
                    row_groups.append((current_start, current_start, current_rows))
                state = "baseline"
                continue
            if _is_difference_header(values):
                state = "difference"
                continue
            if state == "baseline" and current_rows is not None:
                current_rows.append((row_number, values))
                start, _, rows = row_groups[-1]
                row_groups[-1] = (start, row_number, rows)
        global_offset += worksheet.max_row + 1000

    warnings: list[str] = []
    groups: list[BaselineGroup] = []
    for index, (start, end, rows) in enumerate(row_groups, start=1):
        roots: list[BaselineRoot] = []
        current_root: BaselineRoot | None = None
        current_process = "SMT"
        for row_number, values in rows:
            texts = [_clean(value) for value in values]
            process = next(
                (
                    text.upper()
                    for text in texts
                    if text.upper() in {"SMT", "ASSY", "DIP"}
                ),
                "",
            )
            if process:
                current_process = process
            semi_finished = next(
                (text for text in texts if SEMI_RE.fullmatch(text)), ""
            )
            if semi_finished:
                current_root = BaselineRoot(
                    semi_finished=semi_finished,
                    process=current_process,
                    source_items=[],
                    source_row=row_number,
                )
                roots.append(current_root)
            item = _parse_source_item(values, row_number)
            if item is None or current_root is None:
                continue
            current_root.source_items.append(item)
        roots = [root for root in roots if root.source_items]
        for root in roots:
            ratios: list[int] = []
            pcb_ratios: list[int] = []
            for item in root.source_items:
                position_count = len(item.positions)
                declared = item.declared_quantity
                if (
                    declared is None
                    or position_count <= 0
                    or declared <= 0
                    or declared % position_count
                ):
                    continue
                ratio = declared // position_count
                ratios.append(ratio)
                if "PCB" in item.positions:
                    pcb_ratios.append(ratio)
            if pcb_ratios:
                root.batch_multiplier = Counter(pcb_ratios).most_common(1)[0][0]
            elif ratios:
                candidate, count = Counter(ratios).most_common(1)[0]
                root.batch_multiplier = candidate if candidate > 1 and count >= 3 else 1

            mismatched_rows: list[int] = []
            for item in root.source_items:
                if item.declared_quantity is None:
                    continue
                expected = len(item.positions) * root.batch_multiplier
                if item.declared_quantity != expected:
                    mismatched_rows.append(int(item.source_order))
            if root.batch_multiplier > 1:
                warnings.append(
                    f"基準區塊{index}「{root.semi_finished}」自動辨識為整批 "
                    f"{root.batch_multiplier} 倍用量；輸出已換算為單片位置數。"
                )
            if mismatched_rows:
                preview = "、".join(map(str, mismatched_rows[:5]))
                suffix = "…" if len(mismatched_rows) > 5 else ""
                warnings.append(
                    f"基準區塊{index}「{root.semi_finished}」有 "
                    f"{len(mismatched_rows)} 列單量無法由位置數與批量係數核對"
                    f"（列 {preview}{suffix}）；輸出以位置數為準。"
                )
        if roots:
            groups.append(
                BaselineGroup(
                    index=len(groups) + 1,
                    roots=roots,
                    source_start=start,
                    source_end=end,
                )
            )
    if not groups:
        raise AdaptiveBOMError("基準 BOM 找不到可辨識的半成品與料件區塊。")
    return groups, warnings


def _pair_score(difference: DifferenceGroup, baseline: BaselineGroup) -> int:
    common = difference.family_tokens & baseline.family_tokens
    return len(common) * 100


def pair_groups(
    difference_groups: Sequence[DifferenceGroup],
    baseline_groups: Sequence[BaselineGroup],
) -> list[tuple[DifferenceGroup, BaselineGroup, int]]:
    if len(difference_groups) != len(baseline_groups):
        raise AdaptiveBOMError(
            f"自動切割後，差異表有 {len(difference_groups)} 組工單，"
            f"基準 BOM 有 {len(baseline_groups)} 組；數量不同，為避免錯配已停止。"
        )

    direct = [
        (difference, baseline, _pair_score(difference, baseline))
        for difference, baseline in zip(difference_groups, baseline_groups)
    ]
    if all(score > 0 for _, _, score in direct):
        return direct

    remaining = list(baseline_groups)
    paired: list[tuple[DifferenceGroup, BaselineGroup, int]] = []
    for difference in difference_groups:
        candidates = sorted(
            (
                (_pair_score(difference, baseline), baseline)
                for baseline in remaining
            ),
            key=lambda item: item[0],
            reverse=True,
        )
        if not candidates or candidates[0][0] <= 0:
            raise AdaptiveBOMError(
                f"{difference.work_order} 找不到系列相符的基準 BOM 區塊。"
            )
        if len(candidates) > 1 and candidates[0][0] == candidates[1][0]:
            raise AdaptiveBOMError(
                f"{difference.work_order} 有兩個同分的基準 BOM 區塊，無法安全配對。"
            )
        score, baseline = candidates[0]
        remaining.remove(baseline)
        paired.append((difference, baseline, score))
    return paired


def _code_score(root: BaselineRoot, code: str) -> int:
    code = re.escape(code.upper())
    pcb_text = " ".join(
        f"{item.part_no} {item.specification}"
        for item in root.source_items
        if "PCB" in item.positions
    ).upper()
    semi = root.semi_finished.upper()
    # 至少兩位連續數字才視為系列型號，避免把 1ZZ4BIC 的「4B」
    # 誤認成 B 版代碼。
    if re.search(rf"\d{{2,}}{code}(?=$|[A-Z0-9_ /-])", pcb_text):
        return 100
    if re.search(rf"\d{{2,}}{code}(?=$|[A-Z0-9_ /-])", semi):
        return 80
    if re.search(rf"[-_/]\d{{2,}}{code}(?:[-_/]|$)", semi):
        return 50
    return 0


def map_sections(
    difference: DifferenceGroup, baseline: BaselineGroup
) -> dict[str, SectionMapping]:
    mappings: dict[str, SectionMapping] = {}
    used_roots: set[int] = set()

    for section in difference.section_order:
        scored = sorted(
            (
                (_code_score(root, section), index)
                for index, root in enumerate(baseline.roots)
                if index not in used_roots
            ),
            reverse=True,
        )
        if scored and scored[0][0] > 0:
            score, root_index = scored[0]
            if len(scored) == 1 or score > scored[1][0]:
                mappings[section] = SectionMapping(
                    section, root_index, "半成品／PCB代碼", score
                )
                used_roots.add(root_index)

    for section in difference.section_order:
        if section in mappings:
            continue
        rule_positions = {
            rule.position for rule in difference.rules_by_section[section]
        }
        scored = sorted(
            (
                (
                    len(root.position_set & rule_positions),
                    index,
                )
                for index, root in enumerate(baseline.roots)
                if index not in used_roots
            ),
            reverse=True,
        )
        if scored and scored[0][0] >= 2:
            overlap, root_index = scored[0]
            if len(scored) == 1 or overlap > scored[1][0]:
                mappings[section] = SectionMapping(
                    section, root_index, "位置重疊", min(70, 30 + overlap)
                )
                used_roots.add(root_index)

    for section in difference.section_order:
        mappings.setdefault(
            section,
            SectionMapping(section, None, "由差異表建立", 20),
        )
    return mappings


def _legacy_rsp_analysis(
    baseline_source: Source, difference_source: Source
) -> AdaptiveAnalysis | None:
    try:
        details = legacy_rsp.analyze_inputs(baseline_source, difference_source)
    except Exception:
        return None
    if details.get("models") != [1, 2, 3, 4, 5, 6, 7]:
        return None
    targets = [
        TargetPlan(
            key=f"AUTO|{row['機種']}",
            work_order="AUTO",
            model=int(row["機種"]),
            model_name=f"第{row['機種']}機種",
            quantity=None,
            sections=tuple(str(row["自動辨識版別"]).split("/")),
            baseline_sections=tuple(str(row["自動辨識版別"]).split("/")),
            inferred_sections=tuple(),
        )
        for row in details["matrix"]
    ]
    return AdaptiveAnalysis(
        mode="legacy_rsp",
        targets=targets,
        warnings=list(details.get("warnings", [])),
        legacy_details=details,
    )


def _legacy_hlg_analysis(
    baseline_source: Source, difference_source: Source
) -> AdaptiveAnalysis | None:
    workbook = _load(baseline_source)
    text = " ".join(
        _clean(value).upper()
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, 30),
            values_only=True,
        )
        for value in row
        if value is not None
    )
    if "HLG-320H" not in text and "9HLG-" not in text:
        return None
    profile_path = Path(__file__).resolve().parent / "profiles" / "hlg_320h.json"
    try:
        variant = hlg_engine.default_variant_from_profile(profile_path)
        hlg_engine.generate_from_xlsx(
            baseline_source,
            difference_source,
            variant=variant,
            profile_source=profile_path,
        )
    except Exception:
        return None
    return AdaptiveAnalysis(
        mode="legacy_hlg",
        targets=[
            TargetPlan(
                key=f"AUTO|{variant}",
                work_order="AUTO",
                model=variant,
                model_name=f"第{variant}機種",
                quantity=None,
                sections=("A",),
                baseline_sections=("A",),
                inferred_sections=tuple(),
            )
        ],
        warnings=[],
    )


def analyze_inputs(
    baseline_source: Source, difference_source: Source
) -> AdaptiveAnalysis:
    difference_groups = parse_difference_groups(difference_source)

    # V5.1 已驗證過的單系列核心優先保留；是否含工單標題不應影響判斷。
    # 多工單檔才進入 V5.2 的自動切割與配對流程。
    if len(difference_groups) == 1:
        legacy = _legacy_rsp_analysis(baseline_source, difference_source)
        if legacy is not None:
            return legacy
        legacy_hlg = _legacy_hlg_analysis(baseline_source, difference_source)
        if legacy_hlg is not None:
            return legacy_hlg

    baseline_groups, warnings = parse_baseline_groups(baseline_source)
    raw_pairs = pair_groups(difference_groups, baseline_groups)
    paired_groups: list[PairedGroup] = []
    targets: list[TargetPlan] = []

    for difference, baseline, score in raw_pairs:
        mappings = map_sections(difference, baseline)
        pair = PairedGroup(difference, baseline, mappings, score)
        paired_groups.append(pair)
        for model, definition in difference.models.items():
            sections = difference.sections_for_model(model)
            if not sections:
                continue
            baseline_sections = tuple(
                section
                for section in sections
                if mappings[section].baseline_root_index is not None
            )
            inferred_sections = tuple(
                section
                for section in sections
                if mappings[section].baseline_root_index is None
            )
            targets.append(
                TargetPlan(
                    key=f"{difference.work_order}|{model}",
                    work_order=difference.work_order,
                    model=model,
                    model_name=definition.name,
                    quantity=definition.quantity,
                    sections=sections,
                    baseline_sections=baseline_sections,
                    inferred_sections=inferred_sections,
                )
            )
            if inferred_sections:
                warnings.append(
                    f"{difference.work_order} 第{model}機種的 "
                    f"{'/'.join(inferred_sections)} 版在基準 BOM 無獨立半成品，"
                    "將完全依差異表建立並推導半成品號。"
                )
    if not targets:
        raise AdaptiveBOMError("自動配對完成，但沒有任何可產生的機種／版別。")
    return AdaptiveAnalysis(
        mode="adaptive",
        targets=targets,
        warnings=list(dict.fromkeys(warnings)),
        paired_groups=paired_groups,
    )


def _build_items(
    root: BaselineRoot | None,
    rules: Sequence[DiffRule],
    model: int,
    *,
    context: str,
) -> tuple[list[OutputItem], list[str]]:
    warnings: list[str] = []
    assignments: OrderedDict[str, PartSpec] = OrderedDict()
    position_order: dict[str, float] = {}

    if root is not None:
        for item in root.source_items:
            part = PartSpec(item.part_no, item.specification)
            for offset, position in enumerate(item.positions):
                if position in assignments:
                    warnings.append(
                        f"{context}：基準位置 {position} 重複，採用後出現的料件。"
                    )
                assignments[position] = part
                position_order[position] = item.source_order + offset / 1000

    by_position: OrderedDict[str, list[DiffRule]] = OrderedDict()
    for rule in rules:
        by_position.setdefault(rule.position, []).append(rule)

    changed_positions: set[str] = set()
    for position, candidates in by_position.items():
        original = assignments.pop(position, None)
        selected = [rule for rule in candidates if model in rule.models]
        unique = {(rule.part_no, rule.specification) for rule in selected}
        if len(unique) > 1:
            raise AdaptiveBOMError(
                f"{context} 位置 {position} 同時選到多個不同料件。"
            )
        if selected:
            chosen = PartSpec(selected[0].part_no, selected[0].specification)
            assignments[position] = chosen
            position_order.setdefault(position, 100000 + selected[0].source_order)
            if original != chosen:
                changed_positions.add(position)
        elif original is not None:
            changed_positions.add(position)

    grouped: OrderedDict[tuple[str, str], dict[str, object]] = OrderedDict()
    for position, part in assignments.items():
        key = _identity(part.part_no, part.specification)
        bucket = grouped.setdefault(
            key,
            {
                "part": part,
                "positions": [],
                "order": position_order.get(position, 999999.0),
            },
        )
        bucket["positions"].append(position)
        bucket["order"] = min(
            float(bucket["order"]), position_order.get(position, 999999.0)
        )

    output: list[OutputItem] = []
    for bucket in grouped.values():
        part = bucket["part"]
        positions = sorted(
            bucket["positions"], key=lambda value: position_order.get(value, 999999)
        )
        # 基準檔的數量欄可能是整批工單用量；BOM 單量一律以單片
        # 的實際位置數為準，PCB 位置也同樣為 1。
        quantity = len(positions)
        output.append(
            OutputItem(
                part_no=part.part_no,
                specification=part.specification,
                quantity=quantity,
                positions=positions,
                source_order=float(bucket["order"]),
                changed_positions=set(positions) & changed_positions,
            )
        )
    output.sort(
        key=lambda item: (
            0 if "PCB" in item.positions else 1,
            item.source_order,
            _part_key(item.part_no),
        )
    )

    occupied: set[str] = set()
    identities: set[tuple[str, str]] = set()
    for item in output:
        identity = _identity(item.part_no, item.specification)
        if identity in identities:
            raise AdaptiveBOMError(f"{context} 相同品號／規格未合併。")
        identities.add(identity)
        for position in item.positions:
            if position in occupied:
                raise AdaptiveBOMError(f"{context} 位置 {position} 重複。")
            occupied.add(position)

    for position, candidates in by_position.items():
        selected = [rule for rule in candidates if model in rule.models]
        actual = next(
            (
                PartSpec(item.part_no, item.specification)
                for item in output
                if position in item.positions
            ),
            None,
        )
        expected = (
            PartSpec(selected[0].part_no, selected[0].specification)
            if selected
            else None
        )
        if actual != expected:
            raise AdaptiveBOMError(
                f"{context} 位置 {position} 選料驗證失敗。"
            )
    return output, warnings


def _derive_semi_finished(
    root: BaselineRoot | None,
    items: Sequence[OutputItem],
    model_name: str,
    section: str,
) -> str:
    if root is not None and root.semi_finished:
        return root.semi_finished
    pcb = next((item for item in items if "PCB" in item.positions), None)
    if pcb is not None:
        match = re.match(r"^1ZZ\d*(.+)$", pcb.part_no, re.IGNORECASE)
        if match:
            return f"9{match.group(1)}-P"
    tokens = sorted(_family_tokens(model_name))
    if tokens:
        prefix, number = tokens[0]
        return f"9{prefix}-{number}{section}-P"
    return f"9BOM-{section}-P"


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", value).strip() or "BOM"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _format_positions(positions: Sequence[str], max_per_line: int = 4) -> str:
    lines = [
        "   ".join(positions[index : index + max_per_line])
        for index in range(0, len(positions), max_per_line)
    ]
    return "\n".join(lines)


def _write_output_workbook(sheets: Sequence[GeneratedSheet]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    thin_gray = Side(style="thin", color="FFD9D9D9")
    medium_black = Side(style="medium", color="FF000000")
    header_fill = PatternFill("solid", fgColor="FFEAF2F8")

    for generated in sheets:
        worksheet = workbook.create_sheet(generated.sheet_name)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A5"
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.4,
            bottom=0.4,
            header=0.2,
            footer=0.2,
        )
        worksheet.print_title_rows = "1:4"

        for column, width in {
            "A": 10,
            "B": 22,
            "C": 26,
            "D": 47,
            "E": 9,
            "F": 40,
        }.items():
            worksheet.column_dimensions[column].width = width

        worksheet.merge_cells("A1:F1")
        worksheet["A1"] = "BOM清單"
        worksheet["A1"].font = Font(name="Arial", size=16)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="top")
        worksheet.row_dimensions[1].height = 24

        worksheet.merge_cells("A2:F2")
        quantity_text = (
            f"({generated.order_quantity})"
            if generated.order_quantity is not None
            else ""
        )
        worksheet["A2"] = (
            f"工單/機種：{generated.work_order}      "
            f"{generated.model_name}        {quantity_text}"
        )
        worksheet["A2"].font = Font(name="DFKai-SB", size=14)
        worksheet["A2"].alignment = Alignment(horizontal="left", vertical="top")
        worksheet.row_dimensions[2].height = 24

        headers = ["製程段", "半成品", "品號", "規格", "單量", "位置"]
        separators = [
            ".==",
            ".=========",
            ".==============",
            ".===============================",
            ".==",
            ".=================",
        ]
        for column, (header, separator) in enumerate(
            zip(headers, separators), start=1
        ):
            header_cell = worksheet.cell(3, column, header)
            header_cell.font = Font(name="DFKai-SB", size=10, bold=True)
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            separator_cell = worksheet.cell(4, column, separator)
            separator_cell.font = Font(name="MingLiu", size=10)

        current_row = 5
        for index, item in enumerate(generated.items):
            worksheet.cell(current_row, 1, generated.process if index == 0 else None)
            worksheet.cell(
                current_row, 2, generated.semi_finished if index == 0 else None
            )
            worksheet.cell(current_row, 3, item.part_no)
            worksheet.cell(current_row, 4, item.specification)
            worksheet.cell(current_row, 5, item.quantity)
            worksheet.cell(current_row, 6, _format_positions(item.positions))

            for column in range(1, 7):
                cell = worksheet.cell(current_row, column)
                cell.font = Font(name="PMingLiu", size=10)
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=column in (4, 6),
                )
                cell.border = Border(bottom=thin_gray)
            worksheet.cell(current_row, 5).alignment = Alignment(
                horizontal="right", vertical="top"
            )
            worksheet.cell(current_row, 5).number_format = "0"
            if index == 0 and "PCB" in item.positions:
                worksheet.cell(current_row, 3).font = Font(
                    name="PMingLiu", size=10, bold=True, italic=True
                )
                worksheet.cell(current_row, 4).font = Font(
                    name="PMingLiu", size=10, bold=True, italic=True
                )
            if item.changed_positions:
                worksheet.cell(current_row, 6).font = Font(
                    name="MingLiu", size=9, color="FFC00000"
                )
            else:
                worksheet.cell(current_row, 6).font = Font(name="MingLiu", size=9)
            line_count = max(
                1, str(worksheet.cell(current_row, 6).value or "").count("\n") + 1
            )
            worksheet.row_dimensions[current_row].height = max(15, 13.5 * line_count)
            current_row += 1

        first_total_row = 6 if generated.items and "PCB" in generated.items[0].positions else 5
        if current_row - 1 >= first_total_row:
            worksheet.cell(
                current_row, 5, f"=SUM(E{first_total_row}:E{current_row - 1})"
            )
        else:
            worksheet.cell(current_row, 5, 0)
        worksheet.cell(current_row, 5).number_format = "0"
        worksheet.cell(current_row, 6, "End")
        for column in range(1, 7):
            worksheet.cell(current_row, column).border = Border(bottom=medium_black)
        worksheet.cell(current_row, 5).alignment = Alignment(horizontal="right")
        worksheet.cell(current_row, 6).alignment = Alignment(horizontal="right")
        worksheet.auto_filter.ref = f"A3:F{current_row - 1}"
        worksheet.print_area = f"A1:F{current_row}"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _wrap_legacy_rsp(
    baseline_source: Source,
    difference_source: Source,
    selected_targets: Sequence[str],
    project_name: str,
) -> GenerationResult:
    models = [int(key.split("|")[-1]) for key in selected_targets]
    legacy = legacy_rsp.generate_bom(
        baseline_source,
        difference_source,
        models=models,
        project_name=project_name,
        apply_calibration=True,
    )
    sheets: list[GeneratedSheet] = []
    for sheet in legacy.sheets:
        items = [
            OutputItem(
                part_no=item.part_no,
                specification=item.specification,
                quantity=item.quantity,
                positions=list(item.positions),
                source_order=item.source_order,
                changed_positions=set(item.changed_positions),
            )
            for item in sheet.items
        ]
        sheets.append(
            GeneratedSheet(
                target_key=f"AUTO|{sheet.model}",
                work_order_group="AUTO",
                model=sheet.model,
                section=sheet.code,
                sheet_name=sheet.sheet_name,
                model_name=sheet.model_name,
                work_order=sheet.work_order,
                order_quantity=None,
                semi_finished=sheet.semi_finished,
                process=sheet.process,
                items=items,
            )
        )
    return GenerationResult(
        workbook_bytes=legacy.workbook_bytes,
        sheets=sheets,
        warnings=legacy.warnings,
        mode="legacy_rsp",
        target_keys=tuple(selected_targets),
    )


def _wrap_legacy_hlg(
    baseline_source: Source,
    difference_source: Source,
    selected_targets: Sequence[str],
) -> GenerationResult:
    profile_path = Path(__file__).resolve().parent / "profiles" / "hlg_320h.json"
    variant = int(selected_targets[0].split("|")[-1])
    legacy = hlg_engine.generate_from_xlsx(
        baseline_source,
        difference_source,
        variant=variant,
        profile_source=profile_path,
    )
    source_sheet = legacy.sheet
    items = [
        OutputItem(
            part_no=item.part_no,
            specification=item.specification,
            quantity=item.quantity,
            positions=list(item.positions),
            source_order=item.source_order,
            changed_positions=set(item.changed_positions),
        )
        for item in source_sheet.items
    ]
    sheet = GeneratedSheet(
        target_key=selected_targets[0],
        work_order_group="AUTO",
        model=variant,
        section="A",
        sheet_name=source_sheet.sheet_name,
        model_name=source_sheet.model_name,
        work_order=source_sheet.work_order,
        order_quantity=None,
        semi_finished=source_sheet.semi_finished,
        process=source_sheet.process,
        items=items,
    )
    return GenerationResult(
        workbook_bytes=legacy.workbook_bytes,
        sheets=[sheet],
        warnings=legacy.warnings,
        mode="legacy_hlg",
        target_keys=tuple(selected_targets),
    )


def generate_bom(
    baseline_source: Source,
    difference_source: Source,
    *,
    selected_targets: Iterable[str] | None = None,
    project_name: str = "BOM",
) -> GenerationResult:
    analysis = analyze_inputs(baseline_source, difference_source)
    available = {target.key: target for target in analysis.targets}
    requested = tuple(
        dict.fromkeys(selected_targets if selected_targets is not None else available)
    )
    if not requested:
        raise AdaptiveBOMError("至少必須選擇一個工單／機種。")
    unknown = [key for key in requested if key not in available]
    if unknown:
        raise AdaptiveBOMError("找不到選取項目：" + ", ".join(unknown))

    if analysis.mode == "legacy_rsp":
        return _wrap_legacy_rsp(
            baseline_source,
            difference_source,
            requested,
            project_name,
        )
    if analysis.mode == "legacy_hlg":
        return _wrap_legacy_hlg(
            baseline_source,
            difference_source,
            requested,
        )

    requested_set = set(requested)
    used_sheet_names: set[str] = set()
    generated_sheets: list[GeneratedSheet] = []
    warnings = list(analysis.warnings)

    for pair in analysis.paired_groups:
        difference = pair.difference
        for model, definition in difference.models.items():
            target_key = f"{difference.work_order}|{model}"
            if target_key not in requested_set:
                continue
            for section in difference.sections_for_model(model):
                mapping = pair.mappings[section]
                root = (
                    pair.baseline.roots[mapping.baseline_root_index]
                    if mapping.baseline_root_index is not None
                    else None
                )
                context = f"{difference.work_order} 第{model}機種 {section}版"
                items, item_warnings = _build_items(
                    root,
                    difference.rules_by_section[section],
                    model,
                    context=context,
                )
                warnings.extend(item_warnings)
                if not items:
                    raise AdaptiveBOMError(f"{context} 沒有任何料件，已停止輸出。")
                semi_finished = _derive_semi_finished(
                    root,
                    items,
                    definition.name,
                    section,
                )
                raw_sheet_name = (
                    f"第{model}機種-{section}版 BOM表"
                    if difference.work_order == "AUTO"
                    else f"{difference.work_order}-第{model}機種-{section}版"
                )
                sheet_name = _safe_sheet_name(raw_sheet_name, used_sheet_names)
                work_order = (
                    f"第{model}機種"
                    if difference.work_order == "AUTO"
                    else difference.work_order
                )
                generated_sheets.append(
                    GeneratedSheet(
                        target_key=target_key,
                        work_order_group=difference.work_order,
                        model=model,
                        section=section,
                        sheet_name=sheet_name,
                        model_name=definition.name,
                        work_order=work_order,
                        order_quantity=definition.quantity,
                        semi_finished=semi_finished,
                        process=root.process if root else "SMT",
                        items=items,
                    )
                )

    expected = sum(
        len(available[key].sections) for key in requested
    )
    if len(generated_sheets) != expected:
        raise AdaptiveBOMError(
            f"工作表數驗證失敗：應為 {expected}，實際為 {len(generated_sheets)}。"
        )
    workbook_bytes = _write_output_workbook(generated_sheets)
    return GenerationResult(
        workbook_bytes=workbook_bytes,
        sheets=generated_sheets,
        warnings=list(dict.fromkeys(warnings)),
        mode="adaptive",
        target_keys=requested,
    )


__all__ = [
    "AdaptiveAnalysis",
    "AdaptiveBOMError",
    "GeneratedSheet",
    "GenerationResult",
    "TargetPlan",
    "analyze_inputs",
    "generate_bom",
    "parse_baseline_groups",
    "parse_difference_groups",
]
