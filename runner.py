from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from adaptive_engine import (
    AdaptiveBOMError,
    GenerationResult,
    analyze_inputs,
    generate_bom,
)


ProgressCallback = Callable[[int, int, str], None]


class RunnerError(ValueError):
    """Raised when inputs or output paths are invalid."""


@dataclass(frozen=True)
class SavedBOM:
    output_path: Path
    family: str
    models: tuple[int, ...]
    target_keys: tuple[str, ...]
    result: GenerationResult

    @property
    def variant(self) -> int | None:
        return self.models[0] if len(self.models) == 1 else None


def _report(
    callback: ProgressCallback | None,
    current: int,
    total: int,
    message: str,
) -> None:
    if callback:
        callback(current, total, message)


def _validate_xlsx(path: str | Path, label: str) -> Path:
    candidate = Path(path).expanduser()
    if not candidate.is_file():
        raise RunnerError(f"{label}不存在：{candidate}")
    if candidate.suffix.lower() != ".xlsx":
        raise RunnerError(f"{label}必須是 .xlsx 格式。")
    try:
        workbook = load_workbook(candidate, read_only=True, data_only=True)
        workbook.close()
    except Exception as exc:
        raise RunnerError(f"{label}不是可讀取的 .xlsx 活頁簿。") from exc
    return candidate.resolve()


def _output_path(path: str | Path) -> Path:
    if not str(path).strip():
        raise RunnerError("輸出檔案不可空白。")
    candidate = Path(path).expanduser()
    if not candidate.suffix:
        candidate = candidate.with_suffix(".xlsx")
    if candidate.suffix.lower() != ".xlsx":
        raise RunnerError("輸出檔案必須是 .xlsx 格式。")
    return candidate.resolve()


def _save_atomically(data: bytes, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            prefix=f".{output_path.stem}-",
            suffix=".xlsx",
            dir=output_path.parent,
            delete=False,
        ) as temporary_file:
            temporary_file.write(data)
            temporary_path = Path(temporary_file.name)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink()


def detect_family(source: str | Path) -> str:
    """Return a descriptive profile without restricting generation to known names."""

    candidate = _validate_xlsx(source, "基準 BOM")
    workbook = load_workbook(candidate, data_only=True, read_only=True)
    text = " ".join(
        str(value).upper()
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, 40),
            max_col=min(worksheet.max_column, 18),
            values_only=True,
        )
        for value in row
        if value is not None
    )
    workbook.close()
    if "RSP-3000" in text or "RSP-2400" in text:
        return "RSP 系列（相容核心）"
    if "HLG-320H" in text or "9HLG-" in text:
        return "HLG 系列（相容核心）"
    return "自動偵測系列／多工單格式"


def _mode_label(mode: str, work_orders: int) -> str:
    if mode == "legacy_rsp":
        return "RSP 系列（V4/V5.1 相容核心）"
    if mode == "legacy_hlg":
        return "HLG 系列（既有驗證核心）"
    return f"自動偵測格式（{work_orders} 個工單群組）"


def run_generation(
    baseline_path: str | Path,
    difference_path: str | Path,
    output_path: str | Path,
    *,
    selected_targets: Iterable[str] | None = None,
    models: Iterable[int] | None = None,
    variant: int | None = None,
    project_name: str | None = None,
    progress: ProgressCallback | None = None,
) -> SavedBOM:
    total_steps = 5
    _report(progress, 1, total_steps, "正在檢查兩份 Excel 與輸出位置……")
    baseline = _validate_xlsx(baseline_path, "基準 BOM")
    difference = _validate_xlsx(difference_path, "差異表")
    output = _output_path(output_path)
    if output in {baseline, difference}:
        raise RunnerError("輸出檔案不可覆蓋任何輸入檔案。")

    try:
        _report(progress, 2, total_steps, "正在自動尋找表頭、工單、機種與製程區……")
        analysis = analyze_inputs(baseline, difference)
        available = {target.key: target for target in analysis.targets}

        requested: tuple[str, ...]
        if selected_targets is not None:
            requested = tuple(dict.fromkeys(str(key) for key in selected_targets))
        elif models is not None or variant is not None:
            selected_models = {
                int(model)
                for model in (
                    models if models is not None else (int(variant),)
                )
            }
            requested = tuple(
                target.key
                for target in analysis.targets
                if target.model in selected_models
            )
        else:
            requested = tuple(available)

        if not requested:
            raise RunnerError("至少必須選擇一個工單／機種。")
        unknown = [key for key in requested if key not in available]
        if unknown:
            raise RunnerError("找不到選取項目：" + "、".join(unknown))

        expected_sheets = sum(len(available[key].sections) for key in requested)
        _report(
            progress,
            3,
            total_steps,
            f"已辨識 {len(requested)} 個目標，正在產生 {expected_sheets} 張 BOM……",
        )
        result = generate_bom(
            baseline,
            difference,
            selected_targets=requested,
            project_name=project_name or baseline.stem,
        )
    except RunnerError:
        raise
    except AdaptiveBOMError as exc:
        raise RunnerError(str(exc)) from exc

    _report(progress, 4, total_steps, "正在核對選料、單量、合併與重複位置……")
    _save_atomically(result.workbook_bytes, output)
    work_orders = len({sheet.work_order_group for sheet in result.sheets})
    family = _mode_label(result.mode, work_orders)
    selected_models = tuple(
        dict.fromkeys(available[key].model for key in requested)
    )
    _report(progress, 5, total_steps, "BOM 已產生完成。")
    return SavedBOM(
        output_path=output,
        family=family,
        models=selected_models,
        target_keys=requested,
        result=result,
    )


__all__ = ["RunnerError", "SavedBOM", "detect_family", "run_generation"]
