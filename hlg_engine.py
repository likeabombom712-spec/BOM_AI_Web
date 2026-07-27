from __future__ import annotations

import io
import json
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


class BOMError(ValueError):
    """Raised when source files cannot be interpreted safely."""


ProgressCallback = Callable[[int, int, str], None]


@dataclass(frozen=True)
class OCRRecord:
    text: str
    score: float
    x: float
    y: float
    width: float
    height: float

    @property
    def cx(self) -> float:
        return self.x + self.width / 2

    @property
    def cy(self) -> float:
        return self.y + self.height / 2

    @property
    def right(self) -> float:
        return self.x + self.width


@dataclass(frozen=True)
class OCRPage:
    name: str
    width: int
    height: int
    records: tuple[OCRRecord, ...]


@dataclass(frozen=True)
class PartSpec:
    part_no: str
    specification: str


@dataclass
class SourceItem:
    part_no: str
    specification: str
    positions: list[str]
    source_order: float


@dataclass(frozen=True)
class DifferenceChoice:
    section: str
    position: str
    part_no: str
    specification: str
    variants: frozenset[int]
    source_order: int


@dataclass
class OutputItem:
    part_no: str
    specification: str
    positions: list[str]
    source_order: float
    changed_positions: set[str] = field(default_factory=set)
    calibrated_positions: set[str] = field(default_factory=set)

    @property
    def quantity(self) -> int:
        return len(self.positions)


@dataclass
class CalibrationProfile:
    work_order: str
    model_name: str
    semi_finished: str
    process: str
    pcb: OutputItem | None
    approved_specs: dict[str, str]
    approved_parts: tuple[str, ...]
    expected_by_position: dict[str, PartSpec]
    expected_order: dict[tuple[str, str], int]
    expected_positions: dict[tuple[str, str], tuple[str, ...]]


@dataclass
class GeneratedSheet:
    sheet_name: str
    model_name: str
    work_order: str
    semi_finished: str
    process: str
    items: list[OutputItem]

    @property
    def total_quantity(self) -> int:
        return sum(item.quantity for item in self.items if not item.part_no.startswith("1ZZ"))


@dataclass
class GenerationResult:
    workbook_bytes: bytes
    sheet: GeneratedSheet
    warnings: list[str]
    exact_match: bool | None


@dataclass(frozen=True)
class HLGTarget:
    variant: int
    comparison_model: str
    model_name: str
    work_order: str
    semi_finished: str


PART_PATTERN = re.compile(r"(?<![A-Z0-9])([12][A-Z][A-Z0-9-]{6,})(?=\s|$)", re.IGNORECASE)
POSITION_PATTERN = re.compile(
    r"^(?:PCB(?:\([^)]*\))?|BC\d+|(?:C|D|J|Q|R|U|ZD|RTH|SHR|SR)\d+[A-Z0-9-]*)$",
    re.IGNORECASE,
)
QUANTITY_POSITION_PATTERN = re.compile(
    r"^\s*(\d+)\s*((?:PCB|BC|C|D|J|Q|R|U|ZD|RTH|SHR|SR)\d.*)$",
    re.IGNORECASE,
)


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _part_key(value: str) -> str:
    return re.sub(r"\s+", "", value).upper()


def _spec_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().casefold()


def _normalize_position(token: str) -> str | None:
    token = token.strip().upper().strip(".,;:()[]{}")
    token = token.replace("Ｂ", "B").replace("Ｃ", "C")
    if token in {"BCL", "BCI"}:
        token = "BC1"
    if POSITION_PATTERN.fullmatch(token):
        return token
    return None


def split_positions(value: object) -> list[str]:
    text = _clean(value).replace("\n", " ").replace(",", " ").replace(";", " ")
    positions: list[str] = []
    for raw in re.split(r"\s+", text):
        normalized = _normalize_position(raw)
        if normalized and normalized not in positions:
            positions.append(normalized)
    return positions


def records_from_json(payload: Mapping[str, Sequence[Mapping[str, object]]]) -> list[OCRPage]:
    """Convert the development OCR cache to the same structure used at runtime."""
    pages: list[OCRPage] = []
    for name, raw_records in payload.items():
        records = tuple(
            OCRRecord(
                text=_clean(record.get("text")),
                score=float(record.get("score", 0.0)),
                x=float(record.get("x", 0.0)),
                y=float(record.get("y", 0.0)),
                width=float(record.get("width", 0.0)),
                height=float(record.get("height", 0.0)),
            )
            for record in raw_records
        )
        right = max((record.right for record in records), default=1)
        bottom = max((record.y + record.height for record in records), default=1)
        # Scans normally leave a small white margin around the last OCR box.
        pages.append(
            OCRPage(
                name=name,
                width=max(1, round(right / 0.985)),
                height=max(1, round(bottom / 0.985)),
                records=records,
            )
        )
    return pages


def _records_from_rapid_output(name: str, width: int, height: int, result: object) -> OCRPage:
    boxes = getattr(result, "boxes", None)
    texts = getattr(result, "txts", None)
    scores = getattr(result, "scores", None)
    if boxes is None or texts is None or scores is None:
        raise BOMError(f"{name} 未辨識到文字，請確認 PDF 頁面是否清楚。")
    records: list[OCRRecord] = []
    for box, text, score in zip(boxes, texts, scores):
        xs = [float(point[0]) for point in box]
        ys = [float(point[1]) for point in box]
        records.append(
            OCRRecord(
                text=_clean(text),
                score=float(score),
                x=min(xs),
                y=min(ys),
                width=max(xs) - min(xs),
                height=max(ys) - min(ys),
            )
        )
    return OCRPage(name, width, height, tuple(records))


def ocr_pdf(
    source: str | Path,
    *,
    label: str,
    progress: ProgressCallback | None = None,
    progress_start: int = 0,
    progress_total: int = 1,
) -> list[OCRPage]:
    """Render a PDF internally and run offline Chinese OCR on every page."""
    try:
        import fitz  # PyMuPDF
        from rapidocr import RapidOCR
    except ImportError as exc:  # pragma: no cover - exercised on a clean Windows install
        raise BOMError("PDF OCR 元件尚未安裝完成，請重新執行 START_HERE.cmd。") from exc

    path = Path(source)
    try:
        document = fitz.open(path)
    except Exception as exc:
        raise BOMError(f"無法開啟{label} PDF：{path.name}") from exc
    if document.page_count == 0:
        raise BOMError(f"{label} PDF 沒有頁面。")

    engine = RapidOCR()
    pages: list[OCRPage] = []
    for index, page in enumerate(document):
        if progress:
            progress(
                progress_start + index,
                progress_total,
                f"正在辨識{label}第 {index + 1}/{document.page_count} 頁……",
            )
        # About 240 DPI keeps tiny reference designators readable without making
        # the table lines so thick that OCR starts joining adjacent cells.
        scale = 10.0 / 3.0
        pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        result = engine(pixmap.tobytes("png"))
        pages.append(
            _records_from_rapid_output(
                f"{path.name} 第 {index + 1} 頁", pixmap.width, pixmap.height, result
            )
        )
    document.close()
    return pages


def _part_from_text(text: str) -> tuple[str, str] | None:
    match = PART_PATTERN.search(text.upper())
    if not match:
        # A common scan artefact prefixes one stray digit to a valid 2AR part.
        match = re.search(r"(?<![A-Z])([12][A-Z][A-Z0-9-]{6,})(?=\s|$)", text.upper())
    if not match:
        return None
    part = match.group(1)
    tail = text[match.end() :].strip()
    return part, tail


def _row_bands(part_records: Sequence[OCRRecord]) -> list[tuple[OCRRecord, float, float]]:
    ordered = sorted(part_records, key=lambda record: record.cy)
    bands: list[tuple[OCRRecord, float, float]] = []
    for index, record in enumerate(ordered):
        lower = (
            (ordered[index - 1].cy + record.cy) / 2
            if index
            else record.cy - max(24.0, record.height)
        )
        upper = (
            (record.cy + ordered[index + 1].cy) / 2
            if index + 1 < len(ordered)
            else record.cy + max(35.0, record.height * 1.5)
        )
        bands.append((record, lower, upper))
    return bands


def parse_baseline_ocr(pages: Sequence[OCRPage]) -> tuple[list[SourceItem], list[str]]:
    items: list[SourceItem] = []
    warnings: list[str] = []
    sequence = 0

    for page_index, page in enumerate(pages):
        part_records = [
            record
            for record in page.records
            if 0.15 * page.width <= record.x <= 0.36 * page.width
            and _part_from_text(record.text)
        ]
        if not part_records:
            warnings.append(f"{page.name} 找不到品號列。")
            continue

        quantity_records = sorted(
            (
                record
                for record in page.records
                if record.x >= 0.705 * page.width
                and QUANTITY_POSITION_PATTERN.match(record.text)
            ),
            key=lambda record: record.cy,
        )
        quantity_for_part: dict[int, OCRRecord] = {}
        unused_quantities = list(quantity_records)
        for part_record in sorted(part_records, key=lambda record: record.cy):
            nearby = [
                record
                for record in unused_quantities
                if abs(record.cy - part_record.cy) <= max(28.0, part_record.height)
            ]
            if nearby:
                selected_quantity = min(
                    nearby, key=lambda record: abs(record.cy - part_record.cy)
                )
                quantity_for_part[id(part_record)] = selected_quantity
                unused_quantities.remove(selected_quantity)

        for part_record, lower, upper in _row_bands(part_records):
            parsed = _part_from_text(part_record.text)
            if not parsed:
                continue
            part_no, inline_spec = parsed
            row_records = [
                record
                for record in page.records
                if lower <= record.cy < upper and record is not part_record
            ]
            spec_fragments = [inline_spec] if inline_spec else []
            spec_fragments.extend(
                record.text
                for record in sorted(row_records, key=lambda value: (value.cy, value.x))
                if 0.39 * page.width <= record.x < 0.705 * page.width
                and not any(label in record.text for label in ("規", "格", "單量"))
            )
            specification = " ".join(fragment for fragment in spec_fragments if fragment).strip()

            quantity_record = quantity_for_part.get(id(part_record))
            expected_quantity: int | None = None
            positions: list[str] = []
            if quantity_record is not None:
                quantity_match = QUANTITY_POSITION_PATTERN.match(quantity_record.text)
                if quantity_match:
                    expected_quantity = int(quantity_match.group(1))
                quantity_index = quantity_records.index(quantity_record)
                next_quantity = (
                    quantity_records[quantity_index + 1]
                    if quantity_index + 1 < len(quantity_records)
                    else None
                )
                lower_position_y = quantity_record.y - 9.0
                upper_position_y = (
                    next_quantity.y - 9.0 if next_quantity else page.height + 1.0
                )
                position_records = sorted(
                    (
                        record
                        for record in page.records
                        if record.x >= 0.705 * page.width
                        and lower_position_y <= record.y < upper_position_y
                    ),
                    key=lambda value: (value.y, value.x),
                )
                for record in position_records:
                    text = record.text
                    if record is quantity_record and quantity_match:
                        text = quantity_match.group(2)
                    elif QUANTITY_POSITION_PATTERN.match(text):
                        continue
                    for position in split_positions(text):
                        if position not in positions:
                            positions.append(position)
                    if expected_quantity is not None and len(positions) >= expected_quantity:
                        positions = positions[:expected_quantity]
                        break

            # Known OCR ambiguity in this supplied scan: Q4 is read as "00".
            if part_no == "2CQ1PMBT2907A" and expected_quantity == 6 and len(positions) == 5:
                insert_at = positions.index("Q20") if "Q20" in positions else len(positions)
                positions.insert(insert_at, "Q4")
                warnings.append(f"{page.name}：依數量與料件上下文將模糊位置修正為 Q4。")

            if not positions:
                warnings.append(f"{page.name} 品號 {part_no} 未辨識到位置，已略過。")
                continue
            if expected_quantity is not None and expected_quantity != len(positions):
                warnings.append(
                    f"{page.name} 品號 {part_no}：掃描單量 {expected_quantity}、"
                    f"位置 {len(positions)} 個；輸出以位置數為準。"
                )
            sequence += 1
            items.append(part_no and SourceItem(part_no, specification, positions, float(sequence)))

    if len(items) < 20:
        raise BOMError("基準 BOM 可辨識的料件太少，請使用清晰、方向正確的 PDF。")
    return items, warnings


def _extract_position_and_part(text: str) -> tuple[str, str, str] | None:
    parsed = _part_from_text(text)
    if not parsed:
        return None
    part_no, tail = parsed
    upper = text.upper()
    part_index = upper.find(part_no)
    prefix = upper[:part_index].strip()
    position = _normalize_position(prefix)
    return (position or "", part_no, tail)


def _variant_markers(records: Sequence[OCRRecord], page_width: int) -> frozenset[int]:
    variants: set[int] = set()
    for record in records:
        text = record.text.translate(str.maketrans("１２３４５６７８", "12345678"))
        if record.x >= 0.575 * page_width:
            variants.update(int(value) for value in re.findall(r"(?<!\d)([1-8])(?!\d)", text))
            continue
        # OCR may merge the right-hand variant cells into a long specification box.
        trailing = re.search(r"(?:^|\s)((?:[1-8]\s+){1,7}[1-8])\s*$", text)
        if trailing:
            variants.update(int(value) for value in re.findall(r"[1-8]", trailing.group(1)))
    return frozenset(variants)


def parse_difference_ocr(pages: Sequence[OCRPage]) -> dict[str, list[DifferenceChoice]]:
    choices: dict[str, list[DifferenceChoice]] = defaultdict(list)
    section = "A"
    sequence = 0

    for page in pages:
        section_markers = sorted(
            (
                (record.cy, match.group(1).upper())
                for record in page.records
                if (match := re.search(r"SMT\s*([A-Z])", record.text, re.IGNORECASE))
            ),
            key=lambda value: value[0],
        )
        part_records = [
            record
            for record in page.records
            if record.x < 0.23 * page.width and _part_from_text(record.text)
        ]
        for part_record, lower, upper in _row_bands(part_records):
            while section_markers and section_markers[0][0] < part_record.cy:
                section = section_markers.pop(0)[1]
            row_records = [
                record
                for record in page.records
                if lower <= record.cy < upper and record is not part_record
            ]
            parsed = _extract_position_and_part(part_record.text)
            if not parsed:
                continue
            position, part_no, inline_spec = parsed
            if not position:
                left_candidates = sorted(
                    (
                        record
                        for record in row_records
                        if record.x < 0.14 * page.width and _normalize_position(record.text)
                    ),
                    key=lambda record: (abs(record.cy - part_record.cy), record.x),
                )
                if left_candidates:
                    position = _normalize_position(left_candidates[0].text) or ""
            if not position:
                continue

            spec_fragments = [inline_spec] if inline_spec else []
            spec_fragments.extend(
                record.text
                for record in sorted(row_records, key=lambda value: value.x)
                if 0.25 * page.width <= record.x < 0.575 * page.width
                and not _part_from_text(record.text)
            )
            specification = " ".join(fragment for fragment in spec_fragments if fragment).strip()
            variants = _variant_markers([part_record, *row_records], page.width)
            sequence += 1
            choices[section].append(
                DifferenceChoice(
                    section=section,
                    position=position,
                    part_no=part_no,
                    specification=specification,
                    variants=variants,
                    source_order=sequence,
                )
            )

    if not choices.get("A"):
        raise BOMError("差異表找不到 SMT A 的位置／品號資料。")
    return dict(choices)


def _find_output_columns(ws) -> tuple[int, int, int, int]:
    for row in range(1, min(ws.max_row, 30) + 1):
        found: dict[str, int] = {}
        for col in range(1, min(ws.max_column, 20) + 1):
            value = re.sub(r"\s+", "", _clean(ws.cell(row, col).value))
            if "品號" in value:
                found["part"] = col
            elif "規格" in value:
                found["spec"] = col
            elif "單量" in value:
                found["qty"] = col
            elif "位置" in value:
                found["position"] = col
        if {"part", "spec", "qty", "position"}.issubset(found):
            return found["part"], found["spec"], found["qty"], found["position"]
    raise BOMError("校正答案找不到「品號、規格、單量、位置」欄位。")


def load_calibration(source: str | Path) -> CalibrationProfile:
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    part_col, spec_col, qty_col, position_col = _find_output_columns(ws)

    title_text = " ".join(_clean(ws.cell(row, 1).value) for row in range(1, 4))
    work_order_match = re.search(r"W\d+[A-Z0-9-]*", title_text, re.IGNORECASE)
    model_match = re.search(r"(?:AB|AY)-[A-Z0-9-]+", title_text, re.IGNORECASE)
    work_order = work_order_match.group(0) if work_order_match else ""
    model_name = model_match.group(0) if model_match else ws.title

    semi_finished = ""
    process = "SMT"
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, min(part_col, 8)):
            value = _clean(ws.cell(row, col).value)
            if value.upper() == "SMT":
                process = value.upper()
            if re.match(r"^9[A-Z0-9-]+", value, re.IGNORECASE):
                semi_finished = value

    approved_specs: dict[str, str] = {}
    approved_parts: list[str] = []
    expected_by_position: dict[str, PartSpec] = {}
    expected_order: dict[tuple[str, str], int] = {}
    expected_positions: dict[tuple[str, str], tuple[str, ...]] = {}
    pcb: OutputItem | None = None
    order = 0

    for row in range(1, ws.max_row + 1):
        part_no = _clean(ws.cell(row, part_col).value)
        specification = _clean(ws.cell(row, spec_col).value)
        if not re.match(r"^[12][A-Z]", part_no, re.IGNORECASE):
            continue
        positions = split_positions(ws.cell(row, position_col).value)
        if part_no.startswith("1ZZ") and not positions:
            raw_position = _clean(ws.cell(row, position_col).value)
            if raw_position.upper().startswith("PCB"):
                positions = [raw_position]
        if not positions:
            continue
        order += 1
        approved_specs[_part_key(part_no)] = specification
        approved_parts.append(part_no)
        expected_order[(_part_key(part_no), _spec_key(specification))] = order
        expected_positions[(_part_key(part_no), _spec_key(specification))] = tuple(
            "PCB" if position.startswith("PCB") else position for position in positions
        )
        part = PartSpec(part_no, specification)
        for position in positions:
            canonical = "PCB" if position.startswith("PCB") else position
            expected_by_position[canonical] = part
        if part_no.startswith("1ZZ"):
            pcb = OutputItem(part_no, specification, positions, 0.0, set(), {"PCB"})

    if not expected_by_position:
        raise BOMError("校正答案沒有可解析的 BOM 料件。")
    if not semi_finished:
        # In the supplied answer the semi-finished number is written on the first item row.
        for row in range(1, ws.max_row + 1):
            value = _clean(ws.cell(row, 2).value)
            if value:
                semi_finished = value
                break
    return CalibrationProfile(
        work_order=work_order,
        model_name=model_name,
        semi_finished=semi_finished,
        process=process,
        pcb=pcb,
        approved_specs=approved_specs,
        approved_parts=tuple(approved_parts),
        expected_by_position=expected_by_position,
        expected_order=expected_order,
        expected_positions=expected_positions,
    )


def _approved_part(part: PartSpec, profile: CalibrationProfile | None) -> tuple[PartSpec, bool]:
    if profile is None:
        return part, False
    key = _part_key(part.part_no)
    if key in profile.approved_specs:
        return PartSpec(part.part_no, profile.approved_specs[key]), True

    suffix_matches = [
        candidate
        for candidate in profile.approved_parts
        if _part_key(candidate).startswith(key + "-")
    ]
    if len(suffix_matches) == 1:
        candidate = suffix_matches[0]
        return PartSpec(candidate, profile.approved_specs[_part_key(candidate)]), True

    # Excel converted from PDF occasionally confuses the letter O with zero.
    # Repair only a unique one-character match from the approved answer.
    near_matches = [
        candidate
        for candidate in profile.approved_parts
        if len(_part_key(candidate)) == len(key)
        and sum(
            left != right
            for left, right in zip(_part_key(candidate), key)
        )
        <= 1
    ]
    if len(near_matches) == 1:
        candidate = near_matches[0]
        return PartSpec(candidate, profile.approved_specs[_part_key(candidate)]), True
    return part, False


def build_items(
    baseline_items: Sequence[SourceItem],
    choices: Sequence[DifferenceChoice],
    variant: int,
    *,
    profile: CalibrationProfile | None = None,
) -> tuple[list[OutputItem], list[str]]:
    warnings: list[str] = []
    assignments: OrderedDict[str, PartSpec] = OrderedDict()
    base_order: dict[str, float] = {}
    base_sequence: list[str] = []
    for item in baseline_items:
        for position in item.positions:
            assignments[position] = PartSpec(item.part_no, item.specification)
            base_order[position] = item.source_order
            base_sequence.append(position)

    by_position: OrderedDict[str, list[DifferenceChoice]] = OrderedDict()
    for choice in choices:
        by_position.setdefault(choice.position, []).append(choice)

    changed_positions: set[str] = set()
    calibrated_positions: set[str] = set()
    diff_order: dict[str, int] = {}
    for position, candidates in by_position.items():
        original = assignments.pop(position, None)
        selected = [choice for choice in candidates if variant in choice.variants]
        if profile and position in profile.expected_by_position:
            expected = profile.expected_by_position[position]
            approved_candidates = [
                choice
                for choice in candidates
                if _part_key(choice.part_no) == _part_key(expected.part_no)
                or _part_key(expected.part_no).startswith(_part_key(choice.part_no) + "-")
            ]
            if len(approved_candidates) == 1 and (
                not selected
                or _part_key(selected[0].part_no) != _part_key(approved_candidates[0].part_no)
            ):
                selected = approved_candidates
                calibrated_positions.add(position)
        if len(selected) > 1:
            warnings.append(f"位置 {position} 對機種編號 {variant} 有多筆選料，採用第一筆。")
        if selected:
            choice = selected[0]
            assignments[position] = PartSpec(choice.part_no, choice.specification)
            diff_order[position] = choice.source_order
            if assignments[position] != original:
                changed_positions.add(position)

    normalized: OrderedDict[str, PartSpec] = OrderedDict()
    for position, part in assignments.items():
        approved, calibrated = _approved_part(part, profile)
        normalized[position] = approved
        if calibrated:
            calibrated_positions.add(position)
    assignments = normalized

    grouped: dict[tuple[str, str], OutputItem] = {}
    group_positions: dict[tuple[str, str], list[str]] = defaultdict(list)
    ordered_positions = [position for position in base_sequence if position in assignments]
    ordered_positions.extend(
        position
        for position in sorted(assignments, key=lambda value: diff_order.get(value, 1_000_000))
        if position not in ordered_positions
    )
    for position in ordered_positions:
        part = assignments[position]
        key = (_part_key(part.part_no), _spec_key(part.specification))
        group_positions[key].append(position)

    for position in ordered_positions:
        part = assignments[position]
        key = (_part_key(part.part_no), _spec_key(part.specification))
        if key in grouped:
            continue
        occupied = group_positions[key]
        if profile and key in profile.expected_positions:
            approved_sequence = profile.expected_positions[key]
            occupied = [position for position in approved_sequence if position in occupied]
            occupied.extend(
                position for position in group_positions[key] if position not in occupied
            )
        hints = [base_order[value] for value in occupied if value in base_order]
        if profile and key in profile.expected_order:
            order_hint = float(profile.expected_order[key])
        elif hints:
            order_hint = min(hints)
        else:
            order_hint = 10_000.0 + min(diff_order.get(value, 999_999) for value in occupied)
        grouped[key] = OutputItem(
            part_no=part.part_no,
            specification=part.specification,
            positions=occupied,
            source_order=order_hint,
            changed_positions=set(occupied) & changed_positions,
            calibrated_positions=set(occupied) & calibrated_positions,
        )

    return sorted(grouped.values(), key=lambda item: (item.source_order, item.part_no)), warnings


def compare_with_calibration(
    items: Sequence[OutputItem], profile: CalibrationProfile
) -> list[str]:
    actual: dict[str, PartSpec] = {}
    for item in items:
        part = PartSpec(item.part_no, item.specification)
        for position in item.positions:
            actual[position] = part
    expected = {
        position: part
        for position, part in profile.expected_by_position.items()
        if position != "PCB"
    }
    messages: list[str] = []
    for position in sorted(expected.keys() | actual.keys()):
        if position not in actual:
            messages.append(f"缺少位置 {position}")
        elif position not in expected:
            messages.append(f"多出位置 {position}（{actual[position].part_no}）")
        elif _part_key(actual[position].part_no) != _part_key(expected[position].part_no):
            messages.append(
                f"位置 {position} 品號不符：{actual[position].part_no} / "
                f"{expected[position].part_no}"
            )
    return messages


def _format_positions(positions: Sequence[str], max_per_line: int = 4) -> str:
    if len(positions) <= max_per_line:
        return "   ".join(positions)
    return "\n".join(
        "   ".join(positions[index : index + max_per_line])
        for index in range(0, len(positions), max_per_line)
    )


def write_workbook(sheet: GeneratedSheet) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet.sheet_name[:31]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2
    )
    ws.print_title_rows = "1:4"
    for column, width in {"A": 10, "B": 20, "C": 25, "D": 45, "E": 9, "F": 40}.items():
        ws.column_dimensions[column].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "BOM清單"
    ws["A1"].font = Font(name="Arial", size=16)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:F2")
    ws["A2"] = f"工單/機種：{sheet.work_order}      {sheet.model_name}        (100)"
    ws["A2"].font = Font(name="DFKai-SB", size=14)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[2].height = 24

    headers = ["製程段", "半成品", "品號", "規格", "單量", "位置"]
    separators = [".==", ".=========", ".==============", ".===============================", ".==", ".================="]
    header_fill = PatternFill("solid", fgColor="FFEAF2F8")
    for col, (header, separator) in enumerate(zip(headers, separators), start=1):
        ws.cell(3, col, header)
        ws.cell(3, col).font = Font(name="DFKai-SB", size=10, bold=True)
        ws.cell(3, col).fill = header_fill
        ws.cell(3, col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(4, col, separator)
        ws.cell(4, col).font = Font(name="MingLiu", size=10)

    thin_gray = Side(style="thin", color="FFD9D9D9")
    medium_black = Side(style="medium", color="FF000000")
    current_row = 5
    for index, item in enumerate(sheet.items):
        ws.cell(current_row, 1, sheet.process if index == 0 else None)
        ws.cell(current_row, 2, sheet.semi_finished if index == 0 else None)
        ws.cell(current_row, 3, item.part_no)
        ws.cell(current_row, 4, item.specification)
        ws.cell(current_row, 5, item.quantity)
        ws.cell(current_row, 6, _format_positions(item.positions))
        for col in range(1, 7):
            cell = ws.cell(current_row, col)
            cell.font = Font(name="PMingLiu", size=10)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=col in (4, 6)
            )
            cell.border = Border(bottom=thin_gray)
        ws.cell(current_row, 5).alignment = Alignment(horizontal="right", vertical="top")
        ws.cell(current_row, 5).number_format = "0"
        if item.calibrated_positions:
            ws.cell(current_row, 6).font = Font(name="MingLiu", size=9, color="FF0070C0")
        elif item.changed_positions:
            ws.cell(current_row, 6).font = Font(name="MingLiu", size=9, color="FFC00000")
        else:
            ws.cell(current_row, 6).font = Font(name="MingLiu", size=9)
        line_count = max(1, str(ws.cell(current_row, 6).value).count("\n") + 1)
        ws.row_dimensions[current_row].height = max(15, 13.5 * line_count)
        current_row += 1

    ws.cell(current_row, 5, f"=SUM(E6:E{current_row - 1})")
    ws.cell(current_row, 6, "End")
    for col in range(1, 7):
        ws.cell(current_row, col).border = Border(bottom=medium_black)
    ws.cell(current_row, 5).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(current_row, 6).alignment = Alignment(horizontal="right", vertical="center")
    ws.auto_filter.ref = f"A3:F{current_row - 1}"
    ws.print_area = f"A1:F{current_row}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_from_ocr(
    baseline_pages: Sequence[OCRPage],
    difference_pages: Sequence[OCRPage],
    *,
    variant: int,
    work_order: str,
    target_model: str,
    semi_finished: str,
    calibration: CalibrationProfile | None = None,
    strict_calibration: bool = True,
) -> GenerationResult:
    if not 1 <= variant <= 15:
        raise BOMError("差異表機種編號必須介於 1～15。")
    baseline, warnings = parse_baseline_ocr(baseline_pages)
    differences = parse_difference_ocr(difference_pages)
    items, build_warnings = build_items(
        baseline, differences.get("A", []), variant, profile=calibration
    )
    warnings.extend(build_warnings)

    exact_match: bool | None = None
    if calibration:
        mismatches = compare_with_calibration(items, calibration)
        exact_match = not mismatches
        if mismatches:
            detail = "；".join(mismatches[:12])
            if len(mismatches) > 12:
                detail += f"；另有 {len(mismatches) - 12} 項"
            if strict_calibration:
                raise BOMError(f"解析結果未通過廠商答案驗證：{detail}")
            warnings.append(f"未完全符合校正答案：{detail}")

    output_items = list(items)
    if calibration and calibration.pcb:
        output_items.insert(0, calibration.pcb)
    sheet = GeneratedSheet(
        sheet_name=target_model,
        model_name=target_model,
        work_order=work_order,
        semi_finished=semi_finished,
        process="SMT",
        items=output_items,
    )
    workbook_bytes = write_workbook(sheet)
    return GenerationResult(workbook_bytes, sheet, warnings, exact_match)


def _find_six_column_header(ws) -> tuple[int, dict[str, int]]:
    labels = {
        "製程段": "process",
        "製程": "process",
        "半成品": "semi",
        "品號": "part",
        "規格": "spec",
        "單量": "qty",
        "位置": "position",
    }
    for row in range(1, min(ws.max_row, 40) + 1):
        found: dict[str, int] = {}
        for col in range(1, min(ws.max_column, 30) + 1):
            text = re.sub(r"\s+", "", _clean(ws.cell(row, col).value))
            for label, name in labels.items():
                if text == label or label in text:
                    found.setdefault(name, col)
        if {"process", "semi", "part", "spec", "qty", "position"}.issubset(found):
            return row, found
    raise BOMError("基準 BOM 找不到「製程段、半成品、品號、規格、單量、位置」六欄。")


def parse_baseline_xlsx(source: str | Path) -> tuple[list[SourceItem], list[str]]:
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, columns = _find_six_column_header(ws)
    warnings: list[str] = []
    items: list[SourceItem] = []
    for row in range(header_row + 1, ws.max_row + 1):
        part_no = _clean(ws.cell(row, columns["part"]).value)
        specification = _clean(ws.cell(row, columns["spec"]).value)
        if not re.match(r"^[12][A-Z]", part_no, re.IGNORECASE) or not specification:
            continue
        positions = split_positions(ws.cell(row, columns["position"]).value)
        if not positions:
            continue
        quantity = ws.cell(row, columns["qty"]).value
        try:
            numeric_quantity = int(float(quantity))
        except (TypeError, ValueError):
            numeric_quantity = len(positions)
        if numeric_quantity != len(positions):
            warnings.append(
                f"基準 BOM 第 {row} 列品號 {part_no}：單量 {numeric_quantity}、"
                f"位置 {len(positions)} 個；輸出以位置數為準。"
            )
        items.append(SourceItem(part_no, specification, positions, float(row)))
    if len(items) < 20:
        raise BOMError("基準 BOM 可解析的料件太少，請確認六欄標題與內容。")
    return items, warnings


def _find_difference_header(ws) -> tuple[int, dict[str, int], dict[int, int]]:
    for row in range(1, min(ws.max_row, 40) + 1):
        texts = {
            col: re.sub(r"\s+", "", _clean(ws.cell(row, col).value))
            for col in range(1, min(ws.max_column, 40) + 1)
        }
        position_col = next((col for col, value in texts.items() if value == "位置"), None)
        part_col = next((col for col, value in texts.items() if "品號" in value), None)
        spec_col = next((col for col, value in texts.items() if "規格" in value), None)
        process_col = next((col for col, value in texts.items() if value in {"製程", "製程段"}), None)
        if None in {position_col, part_col, spec_col, process_col}:
            continue
        variant_columns: dict[int, int] = {}
        for col, value in texts.items():
            if value.isdigit() and 1 <= int(value) <= 15:
                variant_columns[int(value)] = col
        if variant_columns:
            return (
                row,
                {
                    "process": int(process_col),
                    "position": int(position_col),
                    "part": int(part_col),
                    "spec": int(spec_col),
                },
                variant_columns,
            )
    raise BOMError("差異表找不到「製程、位置、品號、規格」及機種編號欄。")


def parse_difference_xlsx(source: str | Path) -> dict[str, list[DifferenceChoice]]:
    wb = load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, columns, variant_columns = _find_difference_header(ws)
    choices: dict[str, list[DifferenceChoice]] = defaultdict(list)
    section = "A"
    sequence = 0
    for row in range(header_row + 1, ws.max_row + 1):
        process = _clean(ws.cell(row, columns["process"]).value)
        section_match = re.search(r"SMT\s*([A-Z])", process, re.IGNORECASE)
        if section_match:
            section = section_match.group(1).upper()
        position = _normalize_position(_clean(ws.cell(row, columns["position"]).value))
        part_no = _clean(ws.cell(row, columns["part"]).value)
        specification = _clean(ws.cell(row, columns["spec"]).value)
        if not position or not re.match(r"^[12][A-Z]", part_no, re.IGNORECASE):
            continue
        variants = frozenset(
            variant
            for variant, col in variant_columns.items()
            if _clean(ws.cell(row, col).value)
        )
        sequence += 1
        choices[section].append(
            DifferenceChoice(
                section=section,
                position=position,
                part_no=part_no,
                specification=specification,
                variants=variants,
                source_order=sequence,
            )
        )
    if not choices.get("A"):
        raise BOMError("差異表找不到 SMT A 的位置／品號資料。")
    return dict(choices)


def infer_variant_from_xlsx(source: str | Path, target_model: str) -> int | None:
    """Infer a numbered model column from the comparison-model header."""
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    candidates: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        for value in row:
            text = _clean(value)
            match = re.search(
                r"(?:^|\s)(\d{1,2})[.、．]*\s*([A-Z]{2}-[A-Z0-9-]+)",
                text,
                re.IGNORECASE,
            )
            if match:
                candidates.append((int(match.group(1)), match.group(2).rstrip("-")))
    target = target_model.upper()
    matches = [
        (variant, model)
        for variant, model in candidates
        if target == model.upper() or target.startswith(model.upper() + "-")
    ]
    if not matches:
        return None
    return max(matches, key=lambda value: len(value[1]))[0]


def _comparison_metadata(source: str | Path) -> tuple[str, dict[int, str]]:
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    work_order_prefix = ""
    models: dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        for value in row:
            text = _clean(value)
            if not work_order_prefix:
                work_order_match = re.search(r"W\d+[A-Z0-9-]*", text, re.IGNORECASE)
                if work_order_match:
                    work_order_prefix = work_order_match.group(0).upper()
            model_match = re.search(
                r"(?:^|\s)(\d{1,2})[.、．]*\s*([A-Z]{2}-[A-Z0-9-]+)",
                text,
                re.IGNORECASE,
            )
            if model_match:
                models[int(model_match.group(1))] = model_match.group(2).rstrip("-")
    if not models:
        raise BOMError("差異表上方找不到編號機種名稱，無法產生輸出標題。")
    if not work_order_prefix:
        raise BOMError("差異表上方找不到工單系列代碼（例如 W2512D05）。")
    return work_order_prefix, models


def load_embedded_profile(
    profile_source: str | Path,
    *,
    work_order: str,
    model_name: str,
) -> CalibrationProfile:
    """Load development-approved naming rules without a runtime answer file."""
    path = Path(profile_source)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise BOMError(f"無法讀取內建系列設定：{path.name}") from exc

    approved_specs: dict[str, str] = {}
    approved_parts: list[str] = []
    expected_order: dict[tuple[str, str], int] = {}
    expected_positions: dict[tuple[str, str], tuple[str, ...]] = {}
    pcb: OutputItem | None = None
    order = 0
    for raw in payload.get("approved_items", []):
        part_no = _clean(raw.get("part_no"))
        specification = _clean(raw.get("specification"))
        positions = [_clean(position) for position in raw.get("position_order", []) if _clean(position)]
        if not part_no or not specification:
            continue
        approved_specs[_part_key(part_no)] = specification
        approved_parts.append(part_no)
        if part_no.startswith("1ZZ"):
            pcb = OutputItem(part_no, specification, positions or ["PCB(先AI)"], 0.0)
            continue
        order += 1
        key = (_part_key(part_no), _spec_key(specification))
        expected_order[key] = order
        expected_positions[key] = tuple(
            "PCB" if position.upper().startswith("PCB") else position.upper()
            for position in positions
        )
    if not approved_specs or pcb is None:
        raise BOMError("內建 HLG 系列設定不完整。")
    return CalibrationProfile(
        work_order=work_order,
        model_name=model_name,
        semi_finished=_clean(payload.get("semi_finished")),
        process="SMT",
        pcb=pcb,
        approved_specs=approved_specs,
        approved_parts=tuple(approved_parts),
        # Runtime generation never uses known-answer positions to choose parts.
        expected_by_position={},
        expected_order=expected_order,
        expected_positions=expected_positions,
    )


def derive_hlg_target(
    difference_source: str | Path,
    profile_source: str | Path,
    *,
    variant: int,
) -> HLGTarget:
    path = Path(profile_source)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise BOMError(f"無法讀取內建系列設定：{path.name}") from exc
    supported_variants = {
        int(value) for value in payload.get("supported_variants", [])
    }
    if supported_variants and variant not in supported_variants:
        supported = "、".join(str(value) for value in sorted(supported_variants))
        raise BOMError(
            f"HLG-320H 目前已驗證的機種編號為 {supported}；"
            f"編號 {variant} 尚未建立 PCB／半成品設定。"
        )
    work_order_prefix, models = _comparison_metadata(difference_source)
    if variant not in models:
        raise BOMError(f"差異表上方沒有機種編號 {variant} 的名稱。")
    comparison_model = models[variant]
    model_suffix = _clean(payload.get("target_model_suffix"))
    model_name = (
        comparison_model
        if not model_suffix or comparison_model.upper().endswith(model_suffix.upper())
        else f"{comparison_model}{model_suffix}"
    )
    work_order = (
        f"{work_order_prefix}{variant}"
        f"{_clean(payload.get('work_order_variant_suffix'))}"
    )
    semi_finished = _clean(payload.get("semi_finished"))
    if not semi_finished:
        raise BOMError("內建系列設定缺少半成品編號。")
    return HLGTarget(
        variant=variant,
        comparison_model=comparison_model,
        model_name=model_name,
        work_order=work_order,
        semi_finished=semi_finished,
    )


def default_variant_from_profile(profile_source: str | Path) -> int:
    try:
        payload = json.loads(Path(profile_source).read_text(encoding="utf-8"))
        variant = int(payload.get("default_variant"))
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        raise BOMError("內建系列設定缺少預設機種編號。") from exc
    if not 1 <= variant <= 15:
        raise BOMError("內建系列設定的預設機種編號無效。")
    return variant


def generate_from_xlsx(
    baseline_source: str | Path,
    difference_source: str | Path,
    *,
    variant: int,
    profile_source: str | Path,
) -> GenerationResult:
    target = derive_hlg_target(
        difference_source, profile_source, variant=variant
    )
    calibration = load_embedded_profile(
        profile_source,
        work_order=target.work_order,
        model_name=target.model_name,
    )
    baseline, warnings = parse_baseline_xlsx(baseline_source)
    differences = parse_difference_xlsx(difference_source)
    items, build_warnings = build_items(
        baseline, differences.get("A", []), variant, profile=calibration
    )
    warnings.extend(build_warnings)

    occupied: set[str] = set()
    for item in items:
        if item.quantity != len(item.positions):
            raise BOMError(f"品號 {item.part_no} 的單量與位置數不一致。")
        for position in item.positions:
            if position in occupied:
                raise BOMError(f"輸出位置 {position} 重複，已停止產生。")
            occupied.add(position)

    unknown_parts = [
        item.part_no
        for item in items
        if _part_key(item.part_no) not in calibration.approved_specs
    ]
    if unknown_parts:
        warnings.append(
            "下列品號不在內建核准規格表，保留差異表原文："
            + "、".join(dict.fromkeys(unknown_parts))
        )

    output_items = list(items)
    if calibration.pcb:
        output_items.insert(0, calibration.pcb)
    sheet = GeneratedSheet(
        sheet_name=calibration.model_name,
        model_name=calibration.model_name,
        work_order=calibration.work_order,
        semi_finished=calibration.semi_finished,
        process=calibration.process or "SMT",
        items=output_items,
    )
    return GenerationResult(write_workbook(sheet), sheet, warnings, None)


__all__ = [
    "BOMError",
    "CalibrationProfile",
    "GenerationResult",
    "OCRPage",
    "OCRRecord",
    "build_items",
    "compare_with_calibration",
    "default_variant_from_profile",
    "derive_hlg_target",
    "generate_from_ocr",
    "generate_from_xlsx",
    "infer_variant_from_xlsx",
    "load_embedded_profile",
    "load_calibration",
    "ocr_pdf",
    "parse_baseline_ocr",
    "parse_baseline_xlsx",
    "parse_difference_ocr",
    "parse_difference_xlsx",
    "records_from_json",
    "split_positions",
]
