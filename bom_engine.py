from __future__ import annotations

import io
import re
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, Iterable, Sequence

from openpyxl import load_workbook

import v4_engine as v4


BOMError = v4.BOMError

SECTION_RE = re.compile(r"^\s*(?:SMT|ASSY)\s*([A-Z])\s*$", re.IGNORECASE)
MODEL_RE = re.compile(r"(?<!\d)(\d{1,2})(?!\d)")
PART_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_./+\-]{3,}$")


@dataclass(frozen=True)
class DifferenceAnalysis:
    """All selectable rows found in the difference workbook."""

    choices_by_section: OrderedDict[str, tuple[v4.DifferenceChoice, ...]]
    models: tuple[int, ...]
    section_order: tuple[str, ...]

    def sections_for_model(self, model: int) -> tuple[str, ...]:
        return tuple(
            section
            for section in self.section_order
            if any(
                model in choice.variants
                for choice in self.choices_by_section.get(section, ())
            )
        )


@dataclass(frozen=True)
class BatchSheet:
    """One generated worksheet plus its machine/section identity."""

    model: int
    code: str
    generated: v4.GeneratedSheet

    @property
    def section(self) -> str:
        return self.generated.section

    @property
    def sheet_name(self) -> str:
        return self.generated.sheet_name

    @property
    def model_name(self) -> str:
        return self.generated.model_name

    @property
    def work_order(self) -> str:
        return self.generated.work_order

    @property
    def semi_finished(self) -> str:
        return self.generated.semi_finished

    @property
    def process(self) -> str:
        return self.generated.process

    @property
    def items(self) -> list[v4.OutputItem]:
        return self.generated.items

    @property
    def total_quantity(self) -> int:
        return self.generated.total_quantity


@dataclass
class GenerationResult:
    workbook_bytes: bytes
    sheets: list[BatchSheet]
    warnings: list[str]
    models: tuple[int, ...]
    model_sections: dict[int, tuple[str, ...]]
    v4_regression_checked: bool


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _load(source: bytes | bytearray | str | Path | BinaryIO):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    return load_workbook(source, data_only=True)


def _looks_like_part(value: str) -> bool:
    text = _clean(value)
    if not text or " " in text or "=" in text:
        return False
    if text.upper() in {"PCB", "END"}:
        return False
    return bool(PART_RE.fullmatch(text)) and any(char.isdigit() for char in text)


def _parse_membership(value: object) -> frozenset[int]:
    if value is None or isinstance(value, bool):
        return frozenset()
    if isinstance(value, int):
        return frozenset({value}) if 0 < value < 100 else frozenset()
    if isinstance(value, float) and value.is_integer():
        number = int(value)
        return frozenset({number}) if 0 < number < 100 else frozenset()
    text = _clean(value)
    if not text or not re.fullmatch(r"[\d\s,，;/；]+", text):
        return frozenset()
    return frozenset(
        number
        for match in MODEL_RE.findall(text)
        if 0 < (number := int(match)) < 100
    )


def parse_difference(
    source: bytes | bytearray | str | Path | BinaryIO,
) -> DifferenceAnalysis:
    """Read dynamic SMT sections and model numbers to the right of spec.

    No board letters and no model range are hard-coded here. A section belongs
    to a model only when at least one row in that section names the model.
    """

    workbook = _load(source)
    choices: OrderedDict[str, list[v4.DifferenceChoice]] = OrderedDict()
    sequence = 0

    for worksheet in workbook.worksheets:
        current_section: str | None = None
        for row in worksheet.iter_rows(values_only=True):
            values = list(row)
            texts = [_clean(value) for value in values]
            section_match = next(
                (
                    match
                    for text in texts
                    if text and (match := SECTION_RE.fullmatch(text))
                ),
                None,
            )
            if section_match:
                current_section = section_match.group(1).upper()
                choices.setdefault(current_section, [])
                continue
            if current_section is None:
                continue

            nonempty = [(index, text) for index, text in enumerate(texts) if text]
            if len(nonempty) < 3:
                continue
            _, position = nonempty[0]
            _, part_no = nonempty[1]
            spec_index, specification = nonempty[2]
            position = position.upper()
            if not v4.POSITION_TOKEN.fullmatch(position):
                continue
            if not _looks_like_part(part_no):
                continue

            model_numbers: set[int] = set()
            # Business rule: machine numbers are read only from cells to the
            # right of the specification cell, never from part/spec text.
            for value in values[spec_index + 1 :]:
                model_numbers.update(_parse_membership(value))
            if not model_numbers:
                continue

            sequence += 1
            choices[current_section].append(
                v4.DifferenceChoice(
                    section=current_section,
                    position=position,
                    part_no=part_no,
                    specification=specification,
                    variants=frozenset(model_numbers),
                    source_order=sequence,
                )
            )

    populated = OrderedDict(
        (section, tuple(section_choices))
        for section, section_choices in choices.items()
        if section_choices
    )
    if not populated:
        raise BOMError("差異表找不到可解析的 SMT 區段與機種數字。")

    models = tuple(
        sorted(
            {
                model
                for section_choices in populated.values()
                for choice in section_choices
                for model in choice.variants
            }
        )
    )
    if not models:
        raise BOMError("差異表規格右側找不到任何機種數字。")
    return DifferenceAnalysis(populated, models, tuple(populated))


def analyze_inputs(
    baseline_source: bytes | bytearray | str | Path | BinaryIO,
    difference_source: bytes | bytearray | str | Path | BinaryIO,
) -> dict[str, object]:
    analysis = parse_difference(difference_source)
    baselines, baseline_warnings = v4.parse_baseline(baseline_source)
    matrix = [
        {
            "機種": model,
            "自動辨識版別": "/".join(analysis.sections_for_model(model)),
            "工作表數": len(analysis.sections_for_model(model)),
        }
        for model in analysis.models
    ]
    return {
        "models": list(analysis.models),
        "sections": list(analysis.section_order),
        "baseline_sections": list(baselines),
        "matrix": matrix,
        "expected_sheet_count": sum(row["工作表數"] for row in matrix),
        "warnings": baseline_warnings,
    }


def _empty_baseline(code: str) -> v4.BaselineSection:
    return v4.BaselineSection(
        code=code,
        semi_finished="",
        process="SMT",
        source_items=[],
    )


def _pcb_first(items: Sequence[v4.OutputItem]) -> list[v4.OutputItem]:
    return sorted(
        items,
        key=lambda item: (
            0 if "PCB" in item.positions else 1,
            item.source_order,
            v4._part_key(item.part_no),
        ),
    )


def _derive_semi_finished(
    code: str,
    items: Sequence[v4.OutputItem],
    baseline: v4.BaselineSection,
) -> str:
    pcb = next((item for item in items if "PCB" in item.positions), None)
    if pcb is not None:
        text = f"{pcb.part_no} {pcb.specification}".upper()
        patterns = (
            # A target such as RSP-3000A1 is still the A board; the trailing
            # digit is a PCB revision, not part of the semi-finished family.
            rf"([A-Z]{{2,}})[-_ ]*(\d+){re.escape(code)}(?=$|[A-Z0-9_ /-])",
            rf"([A-Z]{{2,}})[-_ ]*(\d+)[A-Z](?=$|[A-Z0-9_ /-])",
        )
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return f"9{match.group(1)}-{match.group(2)}{code}-P"

    if baseline.semi_finished:
        current = baseline.semi_finished
        replaced = re.sub(
            r"(?<=\d)[A-Z](?=-P$)", code, current, flags=re.IGNORECASE
        )
        return replaced or current
    return f"9BOM-{code}-P"


def _position_map(items: Sequence[v4.OutputItem]) -> dict[str, v4.PartSpec]:
    result: dict[str, v4.PartSpec] = {}
    for item in items:
        part = v4.PartSpec(item.part_no, item.specification)
        for position in item.positions:
            if position in result:
                raise BOMError(f"輸出位置 {position} 重複出現在兩個料件群組。")
            result[position] = part
    return result


def _expected_choice(
    section: str,
    model: int,
    position: str,
    candidates: Sequence[v4.DifferenceChoice],
    *,
    apply_calibration: bool,
) -> v4.PartSpec | None:
    selected = [choice for choice in candidates if model in choice.variants]
    if len(selected) > 1:
        unique = {(choice.part_no, choice.specification) for choice in selected}
        if len(unique) > 1:
            raise BOMError(
                f"第{model}機種 / {section}版 / {position} 同時對應多個不同料件。"
            )
    part = (
        v4.PartSpec(selected[0].part_no, selected[0].specification)
        if selected
        else None
    )
    if apply_calibration:
        calibrated = v4._calibrated_override(section, model, position)
        if calibrated is not None:
            part = calibrated
        if part is not None:
            part, _ = v4._normalize_approved_part(part, True)
    return part


def _validate_sheet(
    record: BatchSheet,
    choices: Sequence[v4.DifferenceChoice],
    *,
    apply_calibration: bool,
) -> None:
    if not record.items:
        raise BOMError(f"第{record.model}機種 {record.code}版沒有任何料件。")

    positions = _position_map(record.items)
    if "PCB" not in positions:
        raise BOMError(f"第{record.model}機種 {record.code}版缺少 PCB。")

    identities: set[tuple[str, str]] = set()
    for item in record.items:
        if item.quantity != len(item.positions):
            raise BOMError(
                f"{record.sheet_name} 品號 {item.part_no} 的單量與位置數不一致。"
            )
        identity = (v4._part_key(item.part_no), v4._spec_key(item.specification))
        if identity in identities:
            raise BOMError(
                f"{record.sheet_name} 的相同品號／規格未合併：{item.part_no}。"
            )
        identities.add(identity)

    by_position: OrderedDict[str, list[v4.DifferenceChoice]] = OrderedDict()
    for choice in choices:
        by_position.setdefault(choice.position, []).append(choice)
    for position, candidates in by_position.items():
        expected = _expected_choice(
            record.code,
            record.model,
            position,
            candidates,
            apply_calibration=apply_calibration,
        )
        actual = positions.get(position)
        if expected is None:
            if actual is not None:
                raise BOMError(
                    f"第{record.model}機種 {record.code}版位置 {position} 應移除但仍存在。"
                )
            continue
        if actual != expected:
            actual_text = actual.part_no if actual else "缺少"
            raise BOMError(
                f"第{record.model}機種 {record.code}版位置 {position} 選料錯誤："
                f"應為 {expected.part_no}，實際為 {actual_text}。"
            )


def _check_v4_regression(
    baseline_source: bytes | bytearray | str | Path | BinaryIO,
    difference_source: bytes | bytearray | str | Path | BinaryIO,
    records: Sequence[BatchSheet],
    model_sections: dict[int, tuple[str, ...]],
    *,
    apply_calibration: bool,
) -> bool:
    if not apply_calibration:
        return False
    if model_sections.get(3) != v4.TARGET_SECTIONS:
        return False

    legacy = v4.generate_bom(
        baseline_source,
        difference_source,
        variant=3,
        apply_calibration=True,
    )
    current = {
        record.code: record.generated for record in records if record.model == 3
    }
    for legacy_sheet in legacy.sheets:
        current_sheet = current.get(legacy_sheet.section)
        if current_sheet is None:
            raise BOMError(
                f"V4 回歸檢查失敗：第3機種缺少 {legacy_sheet.section} 版。"
            )
        if v4.logical_rows(current_sheet) != v4.logical_rows(legacy_sheet):
            raise BOMError(
                f"V4 回歸檢查失敗：第3機種 {legacy_sheet.section} 版與 V4 核心不一致。"
            )
    return True


def generate_bom(
    baseline_source: bytes | bytearray | str | Path | BinaryIO,
    difference_source: bytes | bytearray | str | Path | BinaryIO,
    *,
    models: Iterable[int] | None = None,
    project_name: str = "BOM",
    apply_calibration: bool = True,
) -> GenerationResult:
    """Generate one or many machines with the verified V4 build function."""

    analysis = parse_difference(difference_source)
    baselines, warnings = v4.parse_baseline(baseline_source)
    requested_source = analysis.models if models is None else models
    requested = tuple(dict.fromkeys(int(model) for model in requested_source))
    if not requested:
        raise BOMError("至少必須選擇一個機種。")
    unknown = [model for model in requested if model not in analysis.models]
    if unknown:
        raise BOMError(
            "差異表中沒有以下機種：" + ", ".join(str(model) for model in unknown)
        )

    project_name = _clean(project_name) or "BOM"
    model_sections = {
        model: analysis.sections_for_model(model) for model in requested
    }
    empty = [model for model, sections in model_sections.items() if not sections]
    if empty:
        raise BOMError(
            "以下機種沒有任何 SMT 區段：" + ", ".join(map(str, empty))
        )

    records: list[BatchSheet] = []
    for model in requested:
        for code in model_sections[model]:
            baseline = baselines.get(code, _empty_baseline(code))
            items, section_warnings = v4.build_section(
                baseline,
                analysis.choices_by_section.get(code, ()),
                model,
                apply_calibration=apply_calibration,
            )
            warnings.extend(
                f"第{model}機種 {code}版：{warning}"
                for warning in section_warnings
            )
            items = _pcb_first(items)
            semi_finished = _derive_semi_finished(code, items, baseline)
            generated = v4.GeneratedSheet(
                section=code,
                sheet_name=f"第{model}機種-{code}版 BOM表",
                model_name=f"{project_name}      {code}版",
                work_order=f"第{model}機種",
                semi_finished=semi_finished,
                process=baseline.process or "SMT",
                items=items,
            )
            record = BatchSheet(model=model, code=code, generated=generated)
            _validate_sheet(
                record,
                analysis.choices_by_section.get(code, ()),
                apply_calibration=apply_calibration,
            )
            records.append(record)

    expected_sheet_count = sum(len(value) for value in model_sections.values())
    if len(records) != expected_sheet_count:
        raise BOMError(
            f"工作表數檢查失敗：應為 {expected_sheet_count}，實際為 {len(records)}。"
        )
    sheet_names = [record.sheet_name for record in records]
    if len(sheet_names) != len(set(sheet_names)):
        raise BOMError("輸出工作表名稱重複。")

    regression_checked = _check_v4_regression(
        baseline_source,
        difference_source,
        records,
        model_sections,
        apply_calibration=apply_calibration,
    )
    workbook_bytes = v4._write_output_workbook(
        [record.generated for record in records]
    )
    return GenerationResult(
        workbook_bytes=workbook_bytes,
        sheets=records,
        warnings=list(dict.fromkeys(warnings)),
        models=requested,
        model_sections=model_sections,
        v4_regression_checked=regression_checked,
    )


def logical_rows(sheet: BatchSheet) -> list[tuple[str, str, int, tuple[str, ...]]]:
    return v4.logical_rows(sheet.generated)


__all__ = [
    "BOMError",
    "BatchSheet",
    "DifferenceAnalysis",
    "GenerationResult",
    "analyze_inputs",
    "generate_bom",
    "logical_rows",
    "parse_difference",
]
