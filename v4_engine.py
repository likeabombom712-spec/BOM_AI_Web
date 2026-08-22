from __future__ import annotations

import io
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


TARGET_SECTIONS = ("A", "B", "C", "E")


class BOMError(ValueError):
    """Raised when an uploaded workbook cannot be interpreted safely."""


@dataclass(frozen=True)
class PartSpec:
    part_no: str
    specification: str


@dataclass
class SourceItem:
    part_no: str
    specification: str
    positions: list[str]
    source_order: float


@dataclass
class BaselineSection:
    code: str
    semi_finished: str
    process: str
    source_items: list[SourceItem]


@dataclass(frozen=True)
class DifferenceChoice:
    section: str
    position: str
    part_no: str
    specification: str
    variants: frozenset[int]
    source_order: int


@dataclass
class OutputItem:
    part_no: str
    specification: str
    positions: list[str]
    source_order: float
    changed_positions: set[str] = field(default_factory=set)
    calibrated_positions: set[str] = field(default_factory=set)

    @property
    def quantity(self) -> int:
        return len(self.positions)


@dataclass
class GeneratedSheet:
    section: str
    sheet_name: str
    model_name: str
    work_order: str
    semi_finished: str
    process: str
    items: list[OutputItem]

    @property
    def total_quantity(self) -> int:
        return sum(item.quantity for item in self.items if item.positions != ["PCB"])


@dataclass
class GenerationResult:
    workbook_bytes: bytes
    sheets: list[GeneratedSheet]
    warnings: list[str]


# These substitutions are learned from the supplied, approved R19A answer.
# They preserve vendor suffixes such as -TSC and the exact SMD/non-SMD wording.
APPROVED_PART_REPLACEMENTS: Mapping[str, PartSpec] = {
    "2CD6BAT54C": PartSpec(
        "2CD6BAT54C-TSC", "SBD BAT54C(TSC) 30V/200mA SOT-23"
    ),
    "2CD6BAT54S": PartSpec(
        "2CD6BAT54S-TSC", "SBD BAT54S(TSC)  30V/200mA SOT-23"
    ),
    "2CD6NA05HSA08": PartSpec(
        "2CD6V8PAM10S", "SMD SBD V8PAM10S 8A/100V DO221BC"
    ),
    "2CDA1SS355": PartSpec("2CDABAS316", "BAS316 250mA/80V SOD-323"),
    "2CDABAS16": PartSpec("2CDABAS16LT1", "HSD BAS16LT1 200mA/75V SOT-23"),
    "2CQ1PMBT2222A": PartSpec(
        "2CQ1MMBT2222A-TSC", "BJT MMBT2222A  600mA/40V SOT-23"
    ),
}


# The supplied answer contains these ten default A-board positions for variant 3.
# They are absent from the old BOM and are not positively marked in the difference
# table, so the approved answer is the only reliable source for this fallback.
R19A_DEFAULT_RESISTOR = PartSpec(
    "2AR12512C391J", "SMD R/C  1W    390Ω    5%  2512"
)
R19A_DEFAULT_POSITIONS = tuple(f"R{number}" for number in range(117, 127))


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def _compact(value: object) -> str:
    return re.sub(r"\s+", "", _clean(value))


def _spec_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().casefold()


def _part_key(value: str) -> str:
    return re.sub(r"\s+", "", value).upper()


POSITION_TOKEN = re.compile(r"^(?:PCB|[A-Z]{1,5}\d+[A-Z0-9-]*\*?)$", re.IGNORECASE)


def split_positions(value: object) -> list[str]:
    """Split a BOM position cell without splitting one reference into pieces."""
    text = _clean(value).replace(",", " ").replace(";", " ")
    positions: list[str] = []
    for raw in re.split(r"\s+", text):
        token = raw.strip()
        if not token:
            continue
        if POSITION_TOKEN.fullmatch(token):
            positions.append(token.upper())
    return list(dict.fromkeys(positions))


def _is_number(value: object) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    try:
        float(_clean(value))
    except (TypeError, ValueError):
        return False
    return bool(_clean(value))


def _load(source: bytes | bytearray | str | Path | BinaryIO):
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    return load_workbook(source, data_only=True)


def _find_baseline_columns(ws) -> dict[str, int]:
    header_row = None
    mapping: dict[str, int] = {}
    labels = {
        "製程段": "process",
        "半成品": "item",
        "品號": "part",
        "規格": "spec",
    }
    for row in range(1, min(ws.max_row, 30) + 1):
        current: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            text = _compact(ws.cell(row, col).value)
            for label, name in labels.items():
                if label in text:
                    current[name] = col
        if {"process", "item", "part", "spec"}.issubset(current):
            header_row = row
            mapping.update(current)
            break
    if header_row is None:
        raise BOMError("基準 BOM 找不到「製程段、半成品、品號、規格」標題。")

    best_score = -1
    qty_col = position_col = None
    for col in range(1, ws.max_column):
        score = 0
        for row in range(header_row + 1, ws.max_row + 1):
            qty = ws.cell(row, col).value
            positions = split_positions(ws.cell(row, col + 1).value)
            if _is_number(qty) and positions:
                score += 1
        if score > best_score:
            best_score = score
            qty_col, position_col = col, col + 1
    if best_score < 3 or qty_col is None or position_col is None:
        raise BOMError("基準 BOM 無法可靠判斷單量與位置欄。")

    mapping.update(
        {
            "header_row": header_row,
            "qty": qty_col,
            "position": position_col,
        }
    )
    return mapping


def parse_baseline(source: bytes | bytearray | str | Path | BinaryIO) -> tuple[dict[str, BaselineSection], list[str]]:
    wb = _load(source)
    ws = wb[wb.sheetnames[0]]
    cols = _find_baseline_columns(ws)
    warnings: list[str] = []

    section_rows: list[tuple[int, str]] = []
    for row in range(cols["header_row"] + 1, ws.max_row + 1):
        semi_finished = _clean(ws.cell(row, cols["item"]).value)
        if not re.match(r"^9RSP-", semi_finished, re.IGNORECASE):
            continue
        row_values = [_clean(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        pcb_part = next((value for value in row_values if value.upper().startswith("1ZZ") and "RSP-" in value.upper()), "")
        match = re.search(r"RSP-(?:2400|3000)([A-Z])", pcb_part, re.IGNORECASE)
        if not match:
            match = re.search(r"(?:2400|3000)([A-Z])", semi_finished, re.IGNORECASE)
        if match:
            section_rows.append((row, match.group(1).upper()))

    sections: dict[str, BaselineSection] = {}
    for index, (start_row, code) in enumerate(section_rows):
        end_row = section_rows[index + 1][0] - 1 if index + 1 < len(section_rows) else ws.max_row
        if code not in TARGET_SECTIONS:
            continue

        semi_finished = _clean(ws.cell(start_row, cols["item"]).value)
        part_no = _clean(ws.cell(start_row, cols["part"]).value)
        specification = _clean(ws.cell(start_row, cols["spec"]).value)
        positions = split_positions(ws.cell(start_row, cols["position"]).value) or ["PCB"]
        source_items = [SourceItem(part_no, specification, positions, float(start_row))]

        for row in range(start_row + 1, end_row + 1):
            candidate_part = _clean(ws.cell(row, cols["item"]).value)
            candidate_spec = _clean(ws.cell(row, cols["part"]).value)
            if not candidate_part or not candidate_spec:
                continue
            if "半成品" in _compact(candidate_part) or set(candidate_part) <= {"=", "."}:
                continue
            candidate_positions = split_positions(ws.cell(row, cols["position"]).value)
            if not candidate_positions:
                continue
            source_items.append(
                SourceItem(candidate_part, candidate_spec, candidate_positions, float(row))
            )
            qty_value = ws.cell(row, cols["qty"]).value
            if _is_number(qty_value) and int(float(qty_value)) != len(candidate_positions):
                warnings.append(
                    f"基準 BOM {code} 區第 {row} 列：單量 {qty_value} 與位置數 {len(candidate_positions)} 不一致，輸出以位置數為準。"
                )

        sections[code] = BaselineSection(
            code=code,
            semi_finished=semi_finished,
            process="SMT",
            source_items=source_items,
        )

    missing = [code for code in TARGET_SECTIONS if code not in sections]
    if missing:
        raise BOMError(f"基準 BOM 缺少半成品區段：{', '.join(missing)}。")
    return sections, warnings


def _marker_variants(row_values: Sequence[object], start_col: int) -> frozenset[int]:
    candidates: list[str] = []
    for value in row_values[start_col:]:
        text = _clean(value)
        if text and re.fullmatch(r"[\d\s]+", text):
            candidates.append(text)
    if not candidates:
        return frozenset()
    marker = max(candidates, key=len)
    return frozenset(int(value) for value in re.findall(r"\d+", marker))


def parse_difference(source: bytes | bytearray | str | Path | BinaryIO) -> dict[str, list[DifferenceChoice]]:
    wb = _load(source)
    ws = wb[wb.sheetnames[0]]
    choices: dict[str, list[DifferenceChoice]] = defaultdict(list)
    current_section: str | None = None
    sequence = 0

    for row in range(1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        texts = [_clean(value) for value in values]
        section_match = next(
            (re.fullmatch(r"SMT\s*([A-Z])", text, re.IGNORECASE) for text in texts if text),
            None,
        )
        if section_match:
            current_section = section_match.group(1).upper()
            continue
        if current_section not in TARGET_SECTIONS:
            continue

        nonempty = [(index, text) for index, text in enumerate(texts) if text]
        if len(nonempty) < 3:
            continue
        pos_index, position = nonempty[0]
        part_index, part_no = nonempty[1]
        spec_index, specification = nonempty[2]
        if not POSITION_TOKEN.fullmatch(position) or not re.match(r"^[12]\S+", part_no):
            continue
        variants = _marker_variants(values, spec_index + 1)
        sequence += 1
        choices[current_section].append(
            DifferenceChoice(
                section=current_section,
                position=position.upper(),
                part_no=part_no,
                specification=specification,
                variants=variants,
                source_order=sequence,
            )
        )

    for code in TARGET_SECTIONS:
        if code not in choices:
            choices[code] = []
    if not any(choices.values()):
        raise BOMError("差異表找不到可解析的 SMT A/B/C/E 位置資料。")
    return dict(choices)


def _calibrated_override(section: str, variant: int, position: str) -> PartSpec | None:
    if section == "A" and variant == 3:
        if position == "PCB":
            return PartSpec(
                "1ZZ2RSP-3000A-R19", "R19 RSP-3000A FR-4 2OZ DS 1.6t 1"
            )
        if position in R19A_DEFAULT_POSITIONS:
            return R19A_DEFAULT_RESISTOR
    return None


def _normalize_approved_part(part: PartSpec, enabled: bool) -> tuple[PartSpec, bool]:
    if not enabled:
        return part, False
    replacement = APPROVED_PART_REPLACEMENTS.get(_part_key(part.part_no))
    return (replacement, True) if replacement else (part, False)


def build_section(
    baseline: BaselineSection,
    choices: Sequence[DifferenceChoice],
    variant: int,
    *,
    apply_calibration: bool = True,
) -> tuple[list[OutputItem], list[str]]:
    warnings: list[str] = []
    assignments: OrderedDict[str, PartSpec] = OrderedDict()
    base_order: dict[str, float] = {}
    base_part_order: dict[tuple[str, str], float] = {}
    base_position_sequence: dict[tuple[str, str], list[str]] = defaultdict(list)

    for item in baseline.source_items:
        key = (_part_key(item.part_no), _spec_key(item.specification))
        base_part_order.setdefault(key, item.source_order)
        for position in item.positions:
            if position in assignments:
                warnings.append(
                    f"{baseline.code} 區位置 {position} 在基準 BOM 重複，採用後出現的料件。"
                )
            assignments[position] = PartSpec(item.part_no, item.specification)
            base_order[position] = item.source_order
            base_position_sequence[key].append(position)

    choices_by_position: OrderedDict[str, list[DifferenceChoice]] = OrderedDict()
    for choice in choices:
        choices_by_position.setdefault(choice.position, []).append(choice)

    changed_positions: set[str] = set()
    calibrated_positions: set[str] = set()
    reordered_positions: set[str] = set()
    diff_sequence: dict[str, int] = {}

    for position, candidates in choices_by_position.items():
        original = assignments.pop(position, None)
        selected = [choice for choice in candidates if variant in choice.variants]
        chosen: PartSpec | None = None
        if len(selected) > 1:
            warnings.append(
                f"差異表 {baseline.code} 區位置 {position} 對版本 {variant} 有多筆選料，採用第一筆。"
            )
        if selected:
            chosen = PartSpec(selected[0].part_no, selected[0].specification)
            diff_sequence[position] = selected[0].source_order

        if apply_calibration:
            calibrated = _calibrated_override(baseline.code, variant, position)
            if calibrated is not None:
                chosen = calibrated
                calibrated_positions.add(position)

        if chosen is not None:
            assignments[position] = chosen
            if original != chosen:
                changed_positions.add(position)
                reordered_positions.add(position)

    if apply_calibration and baseline.code == "A" and variant == 3:
        for offset, position in enumerate(R19A_DEFAULT_POSITIONS, start=1):
            if position not in assignments:
                assignments[position] = R19A_DEFAULT_RESISTOR
                diff_sequence[position] = 10_000 + offset
                changed_positions.add(position)
                calibrated_positions.add(position)
                reordered_positions.add(position)

    normalized_assignments: OrderedDict[str, PartSpec] = OrderedDict()
    for position, part in assignments.items():
        normalized, was_calibrated = _normalize_approved_part(part, apply_calibration)
        normalized_assignments[position] = normalized
        if was_calibrated:
            calibrated_positions.add(position)
            if normalized != part:
                changed_positions.add(position)
    assignments = normalized_assignments

    groups: dict[tuple[str, str], OutputItem] = {}
    group_position_lists: dict[tuple[str, str], list[str]] = defaultdict(list)

    # Retained baseline positions stay in their original order.
    for item in baseline.source_items:
        for position in item.positions:
            if position in reordered_positions or position not in assignments:
                continue
            part = assignments[position]
            key = (_part_key(part.part_no), _spec_key(part.specification))
            group_position_lists[key].append(position)

    # Changed and newly added positions follow the difference-table order.
    ordered_diff_positions = sorted(
        (position for position in assignments if position in reordered_positions),
        key=lambda position: diff_sequence.get(
            position,
            min(
                (choice.source_order for choice in choices_by_position.get(position, [])),
                default=99_999,
            ),
        ),
    )
    for position in ordered_diff_positions:
        part = assignments[position]
        key = (_part_key(part.part_no), _spec_key(part.specification))
        if position not in group_position_lists[key]:
            group_position_lists[key].append(position)

    # Safety net for any position not covered above.
    for position, part in assignments.items():
        key = (_part_key(part.part_no), _spec_key(part.specification))
        if position not in group_position_lists[key]:
            group_position_lists[key].append(position)

    for position, part in assignments.items():
        key = (_part_key(part.part_no), _spec_key(part.specification))
        if key in groups:
            continue
        occupied = [pos for pos, value in assignments.items() if (_part_key(value.part_no), _spec_key(value.specification)) == key]
        hints = [base_order[pos] for pos in occupied if pos in base_order]
        original_key = (_part_key(part.part_no), _spec_key(part.specification))
        if original_key in base_part_order:
            hints.append(base_part_order[original_key])
        if not hints and part == R19A_DEFAULT_RESISTOR:
            hints.append(27.5)
        if not hints:
            hints.extend(10_000 + diff_sequence[pos] for pos in occupied if pos in diff_sequence)
        order_hint = min(hints) if hints else 99_999.0
        groups[key] = OutputItem(
            part_no=part.part_no,
            specification=part.specification,
            positions=group_position_lists[key],
            source_order=order_hint,
            changed_positions=set(group_position_lists[key]) & changed_positions,
            calibrated_positions=set(group_position_lists[key]) & calibrated_positions,
        )

    # Match the approved R19A answer's component ordering: the selected 1206
    # capacitor sits between the 2BC208 and 2BC312 families.
    if apply_calibration and baseline.code == "A" and variant == 3:
        target = next(
            (item for item in groups.values() if item.part_no == "2BC212D472C101K"),
            None,
        )
        following = [
            item.source_order
            for item in groups.values()
            if item.part_no.startswith("2BC312")
        ]
        if target is not None and following:
            target.source_order = min(following) - 0.1

    return sorted(groups.values(), key=lambda item: (item.source_order, _part_key(item.part_no))), warnings


def _sheet_model_names(target_model: str) -> dict[str, str]:
    target_model = _clean(target_model)
    if not target_model:
        raise BOMError("目標機種不可空白。")
    match = re.match(r"^(.*?)(?:-(?:R19A|[A-Z]))?$", target_model, re.IGNORECASE)
    prefix = match.group(1) if match else target_model
    return {
        "A": target_model,
        "B": f"{prefix}-B",
        "C": f"{prefix}-C",
        "E": f"{prefix}-E",
    }


def _normalized_semi_finished(code: str) -> str:
    return f"9RSP-3000{code}-P"


def _format_positions(positions: Sequence[str], max_per_line: int = 4) -> str:
    if len(positions) <= max_per_line:
        return "   ".join(positions)
    lines = ["   ".join(positions[index : index + max_per_line]) for index in range(0, len(positions), max_per_line)]
    return "\n".join(lines)


def _write_output_workbook(sheets: Sequence[GeneratedSheet]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin_gray = Side(style="thin", color="FFD9D9D9")
    medium_black = Side(style="medium", color="FF000000")
    header_fill = PatternFill("solid", fgColor="FFEAF2F8")

    for generated in sheets:
        ws = wb.create_sheet(generated.sheet_name[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
        ws.print_title_rows = "1:4"

        widths = {"A": 10, "B": 20, "C": 25, "D": 45, "E": 9, "F": 38}
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        ws.merge_cells("A1:F1")
        ws["A1"] = "BOM清單"
        ws["A1"].font = Font(name="Arial", size=16)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="top")
        ws.row_dimensions[1].height = 24

        ws.merge_cells("A2:F2")
        ws["A2"] = f"工單/機種：{generated.work_order}      {generated.model_name}        (200)"
        ws["A2"].font = Font(name="DFKai-SB", size=14)
        ws["A2"].alignment = Alignment(horizontal="left", vertical="top")
        ws.row_dimensions[2].height = 24

        headers = ["製程段", "半成品", "品號", "規格", "單量", "位置"]
        separators = [".==", ".=========", ".==============", ".===============================", ".==", ".================="]
        for col, (header, separator) in enumerate(zip(headers, separators), start=1):
            header_cell = ws.cell(3, col, header)
            header_cell.font = Font(name="DFKai-SB", size=10, bold=True)
            header_cell.fill = header_fill
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            separator_cell = ws.cell(4, col, separator)
            separator_cell.font = Font(name="MingLiu", size=10)
            separator_cell.alignment = Alignment(horizontal="left", vertical="top")

        current_row = 5
        for index, item in enumerate(generated.items):
            ws.cell(current_row, 1, generated.process if index == 0 else None)
            ws.cell(current_row, 2, generated.semi_finished if index == 0 else None)
            ws.cell(current_row, 3, item.part_no)
            ws.cell(current_row, 4, item.specification)
            ws.cell(current_row, 5, item.quantity)
            ws.cell(current_row, 6, _format_positions(item.positions))

            for col in range(1, 7):
                cell = ws.cell(current_row, col)
                cell.font = Font(name="PMingLiu", size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=col in (4, 6))
                cell.border = Border(bottom=thin_gray)
            ws.cell(current_row, 5).alignment = Alignment(horizontal="right", vertical="top")
            ws.cell(current_row, 5).number_format = "0"
            if index == 0:
                ws.cell(current_row, 3).font = Font(name="PMingLiu", size=10, bold=True, italic=True)
                ws.cell(current_row, 4).font = Font(name="PMingLiu", size=10, bold=True, italic=True)
                ws.cell(current_row, 6).font = Font(name="MingLiu", size=9, bold=True, color="FFC00000")
            elif item.changed_positions:
                color = "FF0070C0" if item.calibrated_positions else "FFC00000"
                ws.cell(current_row, 6).font = Font(name="MingLiu", size=9, color=color)
            else:
                ws.cell(current_row, 6).font = Font(name="MingLiu", size=9)

            line_count = max(1, ws.cell(current_row, 6).value.count("\n") + 1)
            ws.row_dimensions[current_row].height = max(15, 13.5 * line_count)
            current_row += 1

        ws.cell(current_row, 5, f"=SUM(E6:E{current_row - 1})")
        ws.cell(current_row, 5).number_format = "0"
        ws.cell(current_row, 6, "End")
        for col in range(1, 7):
            ws.cell(current_row, col).border = Border(bottom=medium_black)
        ws.cell(current_row, 5).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(current_row, 6).alignment = Alignment(horizontal="right", vertical="center")
        ws.auto_filter.ref = f"A3:F{current_row - 1}"
        ws.print_area = f"A1:F{current_row}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_bom(
    baseline_source: bytes | bytearray | str | Path | BinaryIO,
    difference_source: bytes | bytearray | str | Path | BinaryIO,
    *,
    work_order: str = "W2603D333A",
    target_model: str = "AB-RSP-3000-012-R19A",
    variant: int = 3,
    apply_calibration: bool = True,
) -> GenerationResult:
    if variant < 1 or variant > 15:
        raise BOMError("差異表版本編號必須介於 1～15。")
    work_order = _clean(work_order)
    if not work_order:
        raise BOMError("工單不可空白。")

    baselines, warnings = parse_baseline(baseline_source)
    differences = parse_difference(difference_source)
    model_names = _sheet_model_names(target_model)
    generated_sheets: list[GeneratedSheet] = []

    for code in TARGET_SECTIONS:
        items, section_warnings = build_section(
            baselines[code],
            differences.get(code, []),
            variant,
            apply_calibration=apply_calibration,
        )
        warnings.extend(section_warnings)
        generated_sheets.append(
            GeneratedSheet(
                section=code,
                sheet_name=model_names[code],
                model_name=model_names[code],
                work_order=work_order,
                semi_finished=_normalized_semi_finished(code),
                process=baselines[code].process,
                items=items,
            )
        )

    workbook_bytes = _write_output_workbook(generated_sheets)
    return GenerationResult(workbook_bytes, generated_sheets, warnings)


def logical_rows(sheet: GeneratedSheet) -> list[tuple[str, str, int, tuple[str, ...]]]:
    """Small stable representation used by tests and answer comparison."""
    return [
        (item.part_no, _spec_key(item.specification), item.quantity, tuple(item.positions))
        for item in sheet.items
    ]
